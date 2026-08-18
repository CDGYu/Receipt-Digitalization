"""XLSX export: ``Receipts``, ``LineItems``, ``Needs Review``, ``Summary`` (§13).

Export is the terminal DISPLAY boundary of the pipeline (the database is the
source of truth; see SPEC §13). Money is ``Decimal`` everywhere on the internal
money path, but a spreadsheet cell is a display value, so here -- and only here
-- each amount is written as ``float(value)`` together with a numeric
``number_format`` so Excel treats it as a number rather than text. This float
conversion is deliberately confined to the cell-writing helpers below; every
aggregate on the ``Summary`` sheet is summed and divided in ``Decimal`` first
and converted once, at the cell. Nothing downstream reads these cells (exports
read from the database, never the sheet).

**Per-receipt metadata.** :func:`export_workbook` takes plain
:class:`~receipts.extract.schema.ReceiptExtraction` objects, which is all the
two data sheets need. The ``Needs Review`` and ``Summary`` sheets additionally
need facts that live *outside* the extraction -- the receipt id, the routed
status and confidence (§12), the review reason/priority, whether any ERROR
finding is still unresolved, and a link to the stored image. Rather than
requiring ORM rows (which would couple export to the database and to a live
session), those arrive through an optional list of :class:`ReceiptExportRow`
running parallel to ``receipts``. A caller with ORM rows builds them in one
comprehension; a caller with nothing keeps the original call form and still gets
every sheet, just with the unknown columns empty. ``has_unresolved_error`` is
the caller's to compute -- typically
``any(f.severity is Severity.ERROR and not f.resolved_by_repair for f in report.findings)``
-- because a report is not otherwise needed here.

**Not every line item is a ledger row.** The ``LineItems`` sheet carries a
receipt's purchases. A row flagged ``is_template_row`` -- a product line the
form pre-prints and this receipt left blank -- was transcribed upstream so
nothing on the paper is lost, but nobody bought it, and an accounting ledger
that lists it is wrong about what was bought. Those rows are dropped here, and
the ``items`` count on ``Receipts`` counts what is left, so the two sheets
cannot disagree about the same receipt. §13.1 calls that column "count of line
items"; here it counts purchases, deliberately -- that clause was written before
``is_template_row`` existed and has no notion of a row that is not a purchase,
and a workbook contradicting itself is worse than a workbook diverging from a
spec clause that predates the field. A receipt whose rows are *all* blank
still gets its ``Receipts`` row, and still reaches ``Needs Review`` if that is
where it was routed: the receipt exists and somebody has to look at it. Only the
purchases are absent.

DEFERRED FOLLOW-UP: §13's docstring mentions streaming with openpyxl's
``write_only`` mode above 5000 receipts to keep memory flat. That mode forbids
random cell access, which is exactly what the §13.5 formatting pass (row fills,
conditional formatting, content-sized column widths) relies on, so it is left
out rather than half-implemented; it needs its own pass that formats while
appending.
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from receipts.score.confidence import ReceiptStatus
from receipts.score.thresholds import REVIEW_THRESHOLD

if TYPE_CHECKING:
    from receipts.extract.schema import LineItem, ReceiptExtraction

# Display formats. Presentation only; they never change a stored value.
_NUMBER_FORMAT = "#,##0.00"  # money and quantity (§13.5)
_PERCENT_FORMAT = "0.0%"  # confidence and rates (§13.1 col Q)
_TEXT_FORMAT = "@"  # keeps leading zeros on card last4 / receipt no.

_RIGHT = Alignment(horizontal="right")
_BOLD = Font(bold=True)
_LINK_FONT = Font(color="FF0563C1", underline="single")

#: Light red, Excel's own "bad" fill: rows with an unresolved ERROR finding.
_ERROR_FILL = PatternFill(fill_type="solid", start_color="FFFFC7CE", end_color="FFFFC7CE")

#: The §12 review threshold, reused as the amber midpoint of the confidence
#: colour scale so anything below it renders red.
_CONFIDENCE_FLOOR = REVIEW_THRESHOLD

#: §13.5: column widths sized to content, capped so one long URL cannot push a
#: column off the screen.
_MAX_COL_WIDTH = 50

_RATE_QUANTUM = Decimal("0.0001")
_CONFIDENCE_QUANTUM = Decimal("0.001")

_RECEIPT_HEADERS = [
    "receipt_id",
    "merchant",
    "buyer",
    "buyer_tax_id",
    "date",
    "currency",
    "subtotal",
    "tax",
    "discount",
    "total",
    "payment_method",
    "receipt_no",
    "card_last4",
    "items",
    "handwritten",
    "confidence",
    "status",
    "image",
]

_LINEITEM_HEADERS = [
    "receipt_id",
    "position",
    "description",
    "qty",
    "unit_price",
    "line_total",
]

#: §13.3: the sheet a human actually opens. Priority leads because the sheet is
#: sorted by it.
_REVIEW_HEADERS = [
    "priority",
    "receipt_id",
    "merchant",
    "buyer",
    "date",
    "total",
    "confidence",
    "reason",
    "image",
]

_SUMMARY_HEADERS = ["metric", "value"]

#: 1-based column index by header name, derived from the header lists above so
#: the two can never disagree. Every cell below addresses its column through one
#: of these rather than through a literal, because inserting a column moves every
#: column to its right -- as ``buyer``/``buyer_tax_id`` just did on ``Receipts``
#: and ``buyer`` on ``Needs Review``. With literals each value would have gone on
#: being written to its old column, under someone else's header, and silently: a
#: spreadsheet cell accepts anything.
_RECEIPT_COL = {name: i for i, name in enumerate(_RECEIPT_HEADERS, start=1)}
_LINEITEM_COL = {name: i for i, name in enumerate(_LINEITEM_HEADERS, start=1)}
_REVIEW_COL = {name: i for i, name in enumerate(_REVIEW_HEADERS, start=1)}


@dataclass(frozen=True)
class ReceiptExportRow:
    """Per-receipt metadata that does not live on ``ReceiptExtraction``.

    Frozen because export must not mutate what it is handed. Every field is
    optional: whatever the caller does not know stays an empty cell rather than
    being invented.

    ``txn_date`` is an ISO ``YYYY-MM-DD`` string, matching
    ``ReceiptMeta.date`` -- so the date range on ``Summary`` and the sort on
    ``Needs Review`` order correctly as plain strings. ``merchant_name`` and
    ``txn_date`` override the extracted values when supplied (§13.1: prefer the
    canonical merchant name and the normalised date over the raw ones).
    """

    receipt_id: str | None = None
    status: ReceiptStatus | None = None
    confidence: Decimal | None = None
    review_reason: str | None = None
    review_priority: int | None = None
    has_unresolved_error: bool = False
    image_url: str | None = None
    merchant_name: str | None = None
    txn_date: str | None = None


def _write_header(ws, headers: list[str]) -> None:
    """Write the bold header row and freeze it so it stays visible on scroll."""
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = _BOLD
    ws.freeze_panes = "A2"


def _num_cell(ws, row: int, col: int, value: Decimal | None) -> None:
    """Write a numeric value (money or qty) at the display boundary.

    Values are ``Decimal`` on the internal path; a cell is terminal display, so
    we convert to ``float`` here (and only here) with a numeric format and the
    right alignment §13.5 asks for. ``None`` is left as an empty cell -- never
    the string "None".
    """
    if value is None:
        return
    cell = ws.cell(row=row, column=col, value=float(value))
    cell.number_format = _NUMBER_FORMAT
    cell.alignment = _RIGHT


def _ratio_cell(ws, row: int, col: int, value: Decimal | None) -> None:
    """Write a 0-1 ratio (confidence, a rate) as a percentage-formatted number.

    The *stored* value stays the raw ratio so the conditional colour scale can
    compare it against the §12 thresholds; only the display is a percentage.
    """
    if value is None:
        return
    cell = ws.cell(row=row, column=col, value=float(value))
    cell.number_format = _PERCENT_FORMAT
    cell.alignment = _RIGHT


def _text_cell(ws, row: int, col: int, value: str | None) -> None:
    """Write a value Excel would otherwise mangle (§13.5) as text.

    The format is set even when the value is ``None`` so a later hand-typed
    ``0042`` in an empty cell keeps its leading zero too.
    """
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = _TEXT_FORMAT


def _link_cell(ws, row: int, col: int, url: str | None) -> None:
    """Write an image link: a real hyperlink plus link styling when we have one."""
    if not url:
        return
    cell = ws.cell(row=row, column=col, value=url)
    cell.hyperlink = url
    cell.font = _LINK_FONT


def _fill_row(ws, row: int, width: int, fill: PatternFill) -> None:
    for col in range(1, width + 1):
        ws.cell(row=row, column=col).fill = fill


def _autosize(ws, width: int) -> None:
    """Size every column to its widest rendered value, capped (§13.5)."""
    for col in range(1, width + 1):
        widest = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                widest = max(widest, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(_MAX_COL_WIDTH, widest + 2)


def _finalize(ws, headers: list[str], *, last_row: int, confidence_col: int | None = None) -> None:
    """Apply the §13.5 sheet-level formatting.

    Autofilter over the whole used range (header included, so the dropdowns
    attach), a 3-colour scale on ``confidence`` that runs red below the §12
    review threshold, content-sized capped column widths, and no sheet
    protection -- reviewers need to edit for reconciliation.
    """
    width = len(headers)
    ws.auto_filter.ref = f"A1:{get_column_letter(width)}{max(last_row, 1)}"
    ws.protection.sheet = False

    if confidence_col is not None and last_row >= 2:
        letter = get_column_letter(confidence_col)
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{last_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="FFF8696B",  # red
                mid_type="num",
                mid_value=float(_CONFIDENCE_FLOOR),
                mid_color="FFFFEB84",  # amber at the 0.60 floor
                end_type="num",
                end_value=1,
                end_color="FF63BE7B",  # green
            ),
        )

    _autosize(ws, width)


def _purchases(receipt: ReceiptExtraction) -> list[LineItem]:
    """The line items that are purchases -- blank pre-printed rows excluded.

    ``LineItem.is_template_row`` marks a product row the form prints and this
    receipt left blank. It is transcribed so nothing on the paper is lost, but
    it is not a purchase, and an accounting ledger that lists one says somebody
    bought something nobody bought. ``validate.rules`` already keeps these rows
    out of the totals it reconciles (see ``_purchased`` there); this is that same
    exclusion, at the display boundary.
    """
    return [item for item in receipt.line_items if not item.is_template_row]


def _merchant_of(receipt: ReceiptExtraction, row: ReceiptExportRow | None) -> str | None:
    if row is not None and row.merchant_name is not None:
        return row.merchant_name
    return receipt.merchant.name


def _date_of(receipt: ReceiptExtraction, row: ReceiptExportRow | None) -> str | None:
    if row is not None and row.txn_date is not None:
        return row.txn_date
    return receipt.receipt.date


def _write_review_sheet(
    ws,
    entries: list[tuple[Any, ReceiptExtraction, ReceiptExportRow]],
) -> None:
    """§13.3: only ``needs_review`` receipts, worst first.

    Sorted by review priority (lower is more urgent) then date; a missing
    priority sorts last rather than pretending to be urgent. With no metadata at
    all nothing qualifies, and the sheet is written header-only.
    """
    _write_header(ws, _REVIEW_HEADERS)

    pending = [
        (receipt_id, receipt, row)
        for receipt_id, receipt, row in entries
        if row.status is ReceiptStatus.NEEDS_REVIEW
    ]
    pending.sort(
        key=lambda entry: (
            entry[2].review_priority if entry[2].review_priority is not None else 1 << 16,
            _date_of(entry[1], entry[2]) or "",
        )
    )

    row_no = 1
    for receipt_id, receipt, row in pending:
        row_no += 1
        ws.cell(row=row_no, column=_REVIEW_COL["priority"], value=row.review_priority)
        ws.cell(row=row_no, column=_REVIEW_COL["receipt_id"], value=receipt_id)
        ws.cell(row=row_no, column=_REVIEW_COL["merchant"], value=_merchant_of(receipt, row))
        _text_cell(ws, row_no, _REVIEW_COL["buyer"], receipt.buyer.name)
        ws.cell(row=row_no, column=_REVIEW_COL["date"], value=_date_of(receipt, row))
        _num_cell(ws, row_no, _REVIEW_COL["total"], receipt.totals.total)
        _ratio_cell(ws, row_no, _REVIEW_COL["confidence"], row.confidence)
        ws.cell(row=row_no, column=_REVIEW_COL["reason"], value=row.review_reason)
        _link_cell(ws, row_no, _REVIEW_COL["image"], row.image_url)
        if row.has_unresolved_error:
            _fill_row(ws, row_no, len(_REVIEW_HEADERS), _ERROR_FILL)

    _finalize(ws, _REVIEW_HEADERS, last_row=row_no, confidence_col=_REVIEW_COL["confidence"])


def _write_summary_sheet(
    ws,
    entries: list[tuple[Any, ReceiptExtraction, ReceiptExportRow]],
) -> None:
    """§13.4: receipt count, date range, auto-approval rate, average confidence,
    count by status, and total value by merchant.

    Every aggregate is computed in ``Decimal`` and divided defensively: an empty
    list (or an export with no status/confidence metadata) leaves the rate and
    average blank rather than dividing by zero or reporting a fabricated 0%.
    """
    _write_header(ws, _SUMMARY_HEADERS)

    dates = sorted(d for _id, receipt, row in entries if (d := _date_of(receipt, row)))
    confidences = [row.confidence for _id, _receipt, row in entries if row.confidence is not None]
    statuses = [row.status for _id, _receipt, row in entries if row.status is not None]

    rate: Decimal | None = None
    if statuses:
        approved = sum(1 for status in statuses if status is ReceiptStatus.AUTO_APPROVED)
        rate = (Decimal(approved) / Decimal(len(statuses))).quantize(
            _RATE_QUANTUM, rounding=ROUND_HALF_UP
        )

    average: Decimal | None = None
    if confidences:
        average = (sum(confidences, Decimal(0)) / Decimal(len(confidences))).quantize(
            _CONFIDENCE_QUANTUM, rounding=ROUND_HALF_UP
        )

    amounts = [
        receipt.totals.total for _id, receipt, _row in entries if receipt.totals.total is not None
    ]
    total_value = sum(amounts, Decimal(0))

    row_no = 2
    ws.cell(row=row_no, column=1, value="receipts")
    ws.cell(row=row_no, column=2, value=len(entries))

    row_no += 1
    ws.cell(row=row_no, column=1, value="date_from")
    ws.cell(row=row_no, column=2, value=dates[0] if dates else None)

    row_no += 1
    ws.cell(row=row_no, column=1, value="date_to")
    ws.cell(row=row_no, column=2, value=dates[-1] if dates else None)

    row_no += 1
    ws.cell(row=row_no, column=1, value="auto_approval_rate")
    _ratio_cell(ws, row_no, 2, rate)

    row_no += 1
    ws.cell(row=row_no, column=1, value="average_confidence")
    _ratio_cell(ws, row_no, 2, average)

    row_no += 1
    ws.cell(row=row_no, column=1, value="total_value")
    _num_cell(ws, row_no, 2, total_value if entries else None)

    # Count by status. Only the statuses actually present are listed, so the
    # sheet says what happened rather than padding with zeroes.
    by_status: dict[str, int] = {}
    for status in statuses:
        by_status[status.value] = by_status.get(status.value, 0) + 1

    row_no += 2  # one blank row separates the sections
    ws.cell(row=row_no, column=1, value="status").font = _BOLD
    ws.cell(row=row_no, column=2, value="count").font = _BOLD
    for name in sorted(by_status):
        row_no += 1
        ws.cell(row=row_no, column=1, value=name)
        ws.cell(row=row_no, column=2, value=by_status[name])

    # Total value by merchant.
    by_merchant: dict[str, Decimal] = {}
    for _id, receipt, row in entries:
        if receipt.totals.total is None:
            continue
        name = _merchant_of(receipt, row) or "(unknown)"
        by_merchant[name] = by_merchant.get(name, Decimal(0)) + receipt.totals.total

    row_no += 2
    ws.cell(row=row_no, column=1, value="merchant").font = _BOLD
    ws.cell(row=row_no, column=2, value="total").font = _BOLD
    for name in sorted(by_merchant):
        row_no += 1
        ws.cell(row=row_no, column=1, value=name)
        _num_cell(ws, row_no, 2, by_merchant[name])

    _finalize(ws, _SUMMARY_HEADERS, last_row=row_no)


def export_workbook(
    receipts: list[ReceiptExtraction],
    out_path: Path,
    *,
    ids: list[str] | None = None,
    rows: list[ReceiptExportRow] | None = None,
    include_review_sheet: bool = True,
    include_summary: bool = True,
) -> Path:
    """Export receipts to the §13 workbook and return the written path.

    ``ids[i]`` supplies the ``receipt_id`` for each receipt when provided,
    falling back to ``rows[i].receipt_id`` and then to the 1-based index --
    ``ids`` wins because it is the more explicit argument. ``rows[i]`` carries
    the per-receipt metadata (status, confidence, review reason/priority, image
    URL) for ``receipts[i]``; see :class:`ReceiptExportRow`. Missing values
    become empty cells -- nothing is invented.

    ``ids`` and ``rows`` must each match ``receipts`` in length; a mismatch
    raises ``ValueError`` before anything is written. ``include_review_sheet``
    and ``include_summary`` drop the ``Needs Review`` / ``Summary`` sheets when
    a caller wants only the data sheets.
    """
    if ids is not None and len(ids) != len(receipts):
        raise ValueError(
            f"ids has {len(ids)} entries but there are {len(receipts)} receipts; "
            "they must be the same length"
        )
    if rows is not None and len(rows) != len(receipts):
        raise ValueError(
            f"rows has {len(rows)} entries but there are {len(receipts)} receipts; "
            "they must be the same length"
        )

    wb = Workbook()

    receipts_ws = wb.active
    receipts_ws.title = "Receipts"
    _write_header(receipts_ws, _RECEIPT_HEADERS)

    lineitems_ws = wb.create_sheet("LineItems")
    _write_header(lineitems_ws, _LINEITEM_HEADERS)

    # (receipt_id, extraction, metadata) triples, reused by the derived sheets so
    # they can never disagree with the Receipts sheet about a receipt.
    entries: list[tuple[Any, ReceiptExtraction, ReceiptExportRow]] = []

    receipt_row = 1  # row 1 is the header on each sheet
    lineitem_row = 1
    for i, receipt in enumerate(receipts):
        row = rows[i] if rows is not None else ReceiptExportRow()
        receipt_id: str | int
        if ids is not None:
            receipt_id = ids[i]
        elif row.receipt_id is not None:
            receipt_id = row.receipt_id
        else:
            receipt_id = i + 1
        entries.append((receipt_id, receipt, row))

        # What the receipt says was bought. The blank pre-printed rows of a BIR
        # form come out here once, so the count below and the LineItems sheet
        # cannot end up disagreeing about the same receipt.
        purchases = _purchases(receipt)

        receipt_row += 1
        receipts_ws.cell(row=receipt_row, column=_RECEIPT_COL["receipt_id"], value=receipt_id)
        receipts_ws.cell(
            row=receipt_row, column=_RECEIPT_COL["merchant"], value=_merchant_of(receipt, row)
        )
        # Who bought, as against the merchant who sold: the "Sold To" block of a
        # BIR sales invoice. A TIN is digits and separators, never a number, so
        # it goes in as text for the same reason card_last4 does.
        _text_cell(receipts_ws, receipt_row, _RECEIPT_COL["buyer"], receipt.buyer.name)
        _text_cell(receipts_ws, receipt_row, _RECEIPT_COL["buyer_tax_id"], receipt.buyer.tax_id)
        receipts_ws.cell(
            row=receipt_row, column=_RECEIPT_COL["date"], value=_date_of(receipt, row)
        )
        receipts_ws.cell(
            row=receipt_row, column=_RECEIPT_COL["currency"], value=receipt.receipt.currency
        )
        _num_cell(receipts_ws, receipt_row, _RECEIPT_COL["subtotal"], receipt.totals.subtotal)
        _num_cell(receipts_ws, receipt_row, _RECEIPT_COL["tax"], receipt.totals.tax)
        _num_cell(receipts_ws, receipt_row, _RECEIPT_COL["discount"], receipt.totals.discount)
        _num_cell(receipts_ws, receipt_row, _RECEIPT_COL["total"], receipt.totals.total)
        receipts_ws.cell(
            row=receipt_row, column=_RECEIPT_COL["payment_method"], value=receipt.payment.method
        )
        # Text format, or Excel eats the leading zeros (§13.5).
        _text_cell(receipts_ws, receipt_row, _RECEIPT_COL["receipt_no"], receipt.receipt.number)
        _text_cell(receipts_ws, receipt_row, _RECEIPT_COL["card_last4"], receipt.payment.card_last4)
        receipts_ws.cell(row=receipt_row, column=_RECEIPT_COL["items"], value=len(purchases))
        receipts_ws.cell(
            row=receipt_row,
            column=_RECEIPT_COL["handwritten"],
            value="Yes" if receipt.meta.is_handwritten else "No",
        )
        _ratio_cell(receipts_ws, receipt_row, _RECEIPT_COL["confidence"], row.confidence)
        receipts_ws.cell(
            row=receipt_row,
            column=_RECEIPT_COL["status"],
            value=row.status.value if row.status is not None else None,
        )
        _link_cell(receipts_ws, receipt_row, _RECEIPT_COL["image"], row.image_url)
        if row.has_unresolved_error:
            _fill_row(receipts_ws, receipt_row, len(_RECEIPT_HEADERS), _ERROR_FILL)

        for item in purchases:
            lineitem_row += 1
            lineitems_ws.cell(
                row=lineitem_row, column=_LINEITEM_COL["receipt_id"], value=receipt_id
            )
            lineitems_ws.cell(
                row=lineitem_row, column=_LINEITEM_COL["position"], value=item.position
            )
            lineitems_ws.cell(
                row=lineitem_row, column=_LINEITEM_COL["description"], value=item.description_raw
            )
            _num_cell(lineitems_ws, lineitem_row, _LINEITEM_COL["qty"], item.qty)
            _num_cell(lineitems_ws, lineitem_row, _LINEITEM_COL["unit_price"], item.unit_price)
            _num_cell(lineitems_ws, lineitem_row, _LINEITEM_COL["line_total"], item.line_total)

    _finalize(
        receipts_ws,
        _RECEIPT_HEADERS,
        last_row=receipt_row,
        confidence_col=_RECEIPT_COL["confidence"],
    )
    _finalize(lineitems_ws, _LINEITEM_HEADERS, last_row=lineitem_row)

    if include_review_sheet:
        _write_review_sheet(wb.create_sheet("Needs Review"), entries)
    if include_summary:
        _write_summary_sheet(wb.create_sheet("Summary"), entries)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
