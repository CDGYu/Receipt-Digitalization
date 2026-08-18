"""Tests for the minimal XLSX export (Receipts + LineItems sheets).

Export is a terminal DISPLAY boundary, so money lands in cells as numeric
floats with a display ``number_format``. These tests reopen the saved workbook
with openpyxl and assert structure (sheets, header/data row counts, frozen
header) and a few spot-checked values, including that a null money field
(discount) stays an empty cell rather than the string "None".
"""

from __future__ import annotations

from decimal import Decimal

import pytest

openpyxl = pytest.importorskip("openpyxl")

from receipts.export import export_workbook  # noqa: E402
from receipts.extract.schema import (  # noqa: E402
    Buyer,
    LineItem,
    Merchant,
    Payment,
    ReceiptExtraction,
    ReceiptMeta,
    Totals,
)


def _receipt_one() -> ReceiptExtraction:
    return ReceiptExtraction(
        merchant=Merchant(name="Total Wine Co"),
        receipt=ReceiptMeta(date="2024-01-15", currency="USD"),
        line_items=[
            LineItem(
                position=1,
                description_raw="TOTAL WINE CO MERLOT",
                qty=Decimal("1"),
                unit_price=Decimal("12.00"),
                line_total=Decimal("12.00"),
            ),
            LineItem(
                position=2,
                description_raw="CHARDONNAY",
                qty=Decimal("2"),
                unit_price=Decimal("5.00"),
                line_total=Decimal("10.00"),
            ),
        ],
        totals=Totals(
            subtotal=Decimal("22.00"),
            tax=Decimal("1.76"),
            discount=Decimal("2.00"),
            total=Decimal("21.76"),
        ),
        payment=Payment(method="card"),
    )


def _receipt_two() -> ReceiptExtraction:
    # discount is None on purpose: it must export as an empty cell, not "None".
    return ReceiptExtraction(
        merchant=Merchant(name="Corner Store"),
        receipt=ReceiptMeta(date="2024-02-20", currency="USD"),
        line_items=[
            LineItem(
                position=1,
                description_raw="MILK 1L",
                qty=Decimal("3"),
                unit_price=Decimal("1.50"),
                line_total=Decimal("4.50"),
            ),
        ],
        totals=Totals(
            subtotal=Decimal("4.50"),
            tax=Decimal("0.36"),
            discount=None,
            total=Decimal("4.86"),
        ),
        payment=Payment(method="cash"),
    )


def test_export_workbook_builds_two_sheets(tmp_path):
    receipts = [_receipt_one(), _receipt_two()]
    out = tmp_path / "book.xlsx"

    returned = export_workbook(receipts, out_path=out, ids=["r1", "r2"])

    assert returned == out
    assert out.exists()

    wb = openpyxl.load_workbook(out)
    # The two data sheets come first; the review/summary sheets that follow are
    # covered by test_all_four_sheets_present_by_default.
    assert wb.sheetnames[:2] == ["Receipts", "LineItems"]


def test_receipts_sheet_rows_and_values(tmp_path):
    receipts = [_receipt_one(), _receipt_two()]
    out = tmp_path / "book.xlsx"
    export_workbook(receipts, out_path=out, ids=["r1", "r2"])

    ws = openpyxl.load_workbook(out)["Receipts"]

    # 1 header row + 2 data rows.
    assert ws.max_row == 3

    # receipt_id column carries the provided ids.
    assert ws["A2"].value == "r1"
    assert ws["A3"].value == "r2"

    # These are column letters, so they move when a column is inserted: buyer
    # and buyer_tax_id went in at C and D, pushing everything from date
    # rightwards by two. total is J, payment_method K, discount I.
    assert ws["J2"].value == pytest.approx(21.76)

    # merchant and payment_method spot checks.
    assert ws["B2"].value == "Total Wine Co"
    assert ws["K3"].value == "cash"

    # a null money field (discount, column I) is an empty cell, not "None".
    assert ws["I3"].value is None


def test_lineitems_sheet_row_count_and_spot_check(tmp_path):
    receipts = [_receipt_one(), _receipt_two()]
    out = tmp_path / "book.xlsx"
    export_workbook(receipts, out_path=out, ids=["r1", "r2"])

    ws = openpyxl.load_workbook(out)["LineItems"]

    # 3 line items total (2 + 1) plus a header row.
    total_items = sum(len(r.line_items) for r in receipts)
    assert total_items == 3
    assert ws.max_row == total_items + 1

    # First data row is r1's first item.
    assert ws["A2"].value == "r1"
    assert ws["C2"].value == "TOTAL WINE CO MERLOT"
    assert ws["D2"].value == pytest.approx(1)
    assert ws["F2"].value == pytest.approx(12.00)

    # Last data row is r2's only item.
    assert ws["A4"].value == "r2"
    assert ws["C4"].value == "MILK 1L"
    assert ws["D4"].value == pytest.approx(3)
    assert ws["F4"].value == pytest.approx(4.50)


def test_headers_are_frozen_on_both_sheets(tmp_path):
    receipts = [_receipt_one(), _receipt_two()]
    out = tmp_path / "book.xlsx"
    export_workbook(receipts, out_path=out, ids=["r1", "r2"])

    wb = openpyxl.load_workbook(out)
    assert wb["Receipts"].freeze_panes == "A2"
    assert wb["LineItems"].freeze_panes == "A2"


def test_export_workbook_rejects_ids_length_mismatch(tmp_path):
    # Fail fast: a short ids list must raise before any cells are written, not
    # IndexError partway through the sheet.
    receipts = [_receipt_one(), _receipt_two()]
    out = tmp_path / "book.xlsx"
    with pytest.raises(ValueError):
        export_workbook(receipts, out_path=out, ids=["only-one"])
    assert not out.exists()


def test_receipt_id_falls_back_to_index_when_ids_absent(tmp_path):
    receipts = [_receipt_one(), _receipt_two()]
    out = tmp_path / "book.xlsx"
    export_workbook(receipts, out_path=out)

    ws = openpyxl.load_workbook(out)["Receipts"]
    # 1-based index fallback.
    assert ws["A2"].value == 1
    assert ws["A3"].value == 2


# --------------------------------------------------------------------------- #
# Needs Review + Summary sheets and the §13.5 formatting.
#
# Per-receipt metadata that does not live on ReceiptExtraction (id, status,
# confidence, review reason/priority, image URL) is supplied through the
# optional parallel ``rows`` argument, so the exporter stays decoupled from the
# ORM and every existing call form keeps working.
# --------------------------------------------------------------------------- #

from receipts.export import ReceiptExportRow  # noqa: E402
from receipts.score.confidence import ReceiptStatus  # noqa: E402


def _simple_receipt(
    merchant: str,
    date: str,
    total: Decimal,
    *,
    number: str | None = None,
    card_last4: str | None = None,
) -> ReceiptExtraction:
    return ReceiptExtraction(
        merchant=Merchant(name=merchant),
        receipt=ReceiptMeta(date=date, currency="USD", number=number),
        line_items=[
            LineItem(
                position=1,
                description_raw="THING",
                qty=Decimal("1"),
                unit_price=total,
                line_total=total,
            )
        ],
        totals=Totals(subtotal=total, tax=Decimal("0.00"), total=total),
        payment=Payment(method="card", card_last4=card_last4),
    )


def _fixture() -> tuple[list[ReceiptExtraction], list[ReceiptExportRow]]:
    """Four receipts: one auto-approved, three needing review.

    Review order must come out priority-then-date: r3 (priority 0), then r4 and
    r2 (both priority 2, ordered by date 2024-01-05 before 2024-02-20).
    """
    receipts = [
        _simple_receipt(
            "Total Wine Co", "2024-01-15", Decimal("21.76"), number="0042", card_last4="0007"
        ),
        _simple_receipt("Corner Store", "2024-02-20", Decimal("4.86")),
        _simple_receipt("Bodega", "2024-03-01", Decimal("99.00")),
        _simple_receipt("Kiosk", "2024-01-05", Decimal("7.50")),
    ]
    rows = [
        ReceiptExportRow(
            receipt_id="r1",
            status=ReceiptStatus.AUTO_APPROVED,
            confidence=Decimal("0.920"),
        ),
        ReceiptExportRow(
            receipt_id="r2",
            status=ReceiptStatus.NEEDS_REVIEW,
            confidence=Decimal("0.700"),
            review_reason="quick verify",
            review_priority=2,
        ),
        ReceiptExportRow(
            receipt_id="r3",
            status=ReceiptStatus.NEEDS_REVIEW,
            confidence=Decimal("0.300"),
            review_reason="urgent: validation errors and total is missing",
            review_priority=0,
            has_unresolved_error=True,
            image_url="https://example.test/img/r3.jpg",
        ),
        ReceiptExportRow(
            receipt_id="r4",
            status=ReceiptStatus.NEEDS_REVIEW,
            confidence=Decimal("0.650"),
            review_reason="quick verify",
            review_priority=2,
        ),
    ]
    return receipts, rows


def _header_col(ws, name: str) -> int:
    """1-based index of the column whose header cell equals ``name``."""
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == name:
            return col
    raise AssertionError(f"{ws.title!r} has no {name!r} column")


def _label_row(ws, label: str) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == label:
            return row
    raise AssertionError(f"{ws.title!r} has no row labelled {label!r}")


def _summary_value(ws, label: str):
    return ws.cell(row=_label_row(ws, label), column=2).value


def _section_pairs(ws, section_label: str) -> dict:
    """The label/value pairs under a section header, up to the next blank row."""
    pairs = {}
    row = _label_row(ws, section_label) + 1
    while row <= ws.max_row and ws.cell(row=row, column=1).value is not None:
        pairs[ws.cell(row=row, column=1).value] = ws.cell(row=row, column=2).value
        row += 1
    return pairs


def _export(tmp_path, *, with_rows: bool = True, **kwargs):
    receipts, rows = _fixture()
    out = tmp_path / "full.xlsx"
    export_workbook(receipts, out_path=out, rows=rows if with_rows else None, **kwargs)
    return openpyxl.load_workbook(out)


def test_all_four_sheets_present_by_default(tmp_path):
    wb = _export(tmp_path)
    assert wb.sheetnames == ["Receipts", "LineItems", "Needs Review", "Summary"]


def test_flags_omit_their_sheets(tmp_path):
    wb = _export(tmp_path, include_review_sheet=False, include_summary=False)
    assert wb.sheetnames == ["Receipts", "LineItems"]

    wb = _export(tmp_path, include_summary=False)
    assert wb.sheetnames == ["Receipts", "LineItems", "Needs Review"]

    wb = _export(tmp_path, include_review_sheet=False)
    assert wb.sheetnames == ["Receipts", "LineItems", "Summary"]


def test_review_sheet_holds_only_needs_review_in_priority_then_date_order(tmp_path):
    ws = _export(tmp_path)["Needs Review"]

    id_col = _header_col(ws, "receipt_id")
    ids = [ws.cell(row=r, column=id_col).value for r in range(2, ws.max_row + 1)]

    # r1 is auto_approved and must not appear at all.
    assert ids == ["r3", "r4", "r2"]


def test_review_sheet_shows_reason_and_image_hyperlink(tmp_path):
    ws = _export(tmp_path)["Needs Review"]

    reason_col = _header_col(ws, "reason")
    assert ws.cell(row=2, column=reason_col).value == (
        "urgent: validation errors and total is missing"
    )
    assert ws.cell(row=3, column=reason_col).value == "quick verify"

    image_col = _header_col(ws, "image")
    linked = ws.cell(row=2, column=image_col)
    assert linked.hyperlink is not None
    assert linked.hyperlink.target == "https://example.test/img/r3.jpg"
    # No URL supplied for r4 -> empty cell, no hyperlink, nothing invented.
    assert ws.cell(row=3, column=image_col).hyperlink is None


def test_review_sheet_is_header_only_without_metadata(tmp_path):
    wb = _export(tmp_path, with_rows=False)
    ws = wb["Needs Review"]
    # Status is unknown without metadata: the sheet still exists, with just the
    # header row, rather than crashing or guessing.
    assert ws.max_row == 1
    assert ws.cell(row=1, column=1).value is not None


def test_summary_reports_counts_dates_rate_and_average(tmp_path):
    ws = _export(tmp_path)["Summary"]

    assert _summary_value(ws, "receipts") == 4
    assert _summary_value(ws, "date_from") == "2024-01-05"
    assert _summary_value(ws, "date_to") == "2024-03-01"
    # 1 of 4 auto-approved.
    assert _summary_value(ws, "auto_approval_rate") == pytest.approx(0.25)
    # (0.920 + 0.700 + 0.300 + 0.650) / 4 = 0.6425 -> 0.643 at 3dp.
    assert _summary_value(ws, "average_confidence") == pytest.approx(0.643)

    by_status = _section_pairs(ws, "status")
    assert by_status == {"auto_approved": 1, "needs_review": 3}

    by_merchant = _section_pairs(ws, "merchant")
    assert by_merchant["Bodega"] == pytest.approx(99.00)
    assert set(by_merchant) == {"Bodega", "Corner Store", "Kiosk", "Total Wine Co"}


def test_summary_survives_an_empty_receipt_list(tmp_path):
    out = tmp_path / "empty.xlsx"
    export_workbook([], out_path=out)

    ws = openpyxl.load_workbook(out)["Summary"]
    assert _summary_value(ws, "receipts") == 0
    # An undefined rate/average is empty, never a fabricated 0% and never a
    # ZeroDivisionError.
    assert _summary_value(ws, "auto_approval_rate") is None
    assert _summary_value(ws, "average_confidence") is None
    assert _summary_value(ws, "date_from") is None


def test_unresolved_error_row_gets_a_light_red_fill(tmp_path):
    ws = _export(tmp_path)["Receipts"]

    # r3 (has_unresolved_error) is the third data row; r1 is clean.
    flagged = ws.cell(row=4, column=1)
    clean = ws.cell(row=2, column=1)
    assert flagged.fill.fill_type == "solid"
    assert flagged.fill.start_color.rgb != clean.fill.start_color.rgb


def test_receipt_no_and_card_last4_are_text_formatted(tmp_path):
    ws = _export(tmp_path)["Receipts"]

    no_col = _header_col(ws, "receipt_no")
    card_col = _header_col(ws, "card_last4")

    assert ws.cell(row=2, column=no_col).value == "0042"
    assert ws.cell(row=2, column=card_col).value == "0007"
    for row in range(2, ws.max_row + 1):
        assert ws.cell(row=row, column=no_col).number_format == "@"
        assert ws.cell(row=row, column=card_col).number_format == "@"


def test_confidence_columns_carry_a_colour_scale_rule(tmp_path):
    from openpyxl.utils import get_column_letter

    wb = _export(tmp_path)
    for name in ("Receipts", "Needs Review"):
        ws = wb[name]
        letter = get_column_letter(_header_col(ws, "confidence"))
        rules = [
            rule
            for fmt in ws.conditional_formatting
            if letter in str(fmt.sqref)
            for rule in fmt.rules
        ]
        assert any(rule.type == "colorScale" for rule in rules), name


def test_autofilter_and_capped_column_widths_on_every_sheet(tmp_path):
    wb = _export(tmp_path)
    for ws in wb.worksheets:
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is not None
        assert ws.protection.sheet is False
        widths = [dim.width for dim in ws.column_dimensions.values() if dim.width]
        assert widths, ws.title
        assert max(widths) <= 50, ws.title


def test_rows_length_mismatch_is_rejected(tmp_path):
    receipts, rows = _fixture()
    out = tmp_path / "book.xlsx"
    with pytest.raises(ValueError):
        export_workbook(receipts, out_path=out, rows=rows[:2])
    assert not out.exists()


def test_ids_take_precedence_over_row_metadata_ids(tmp_path):
    receipts, rows = _fixture()
    out = tmp_path / "book.xlsx"
    export_workbook(receipts, out_path=out, ids=["a", "b", "c", "d"], rows=rows)

    wb = openpyxl.load_workbook(out)
    assert wb["Receipts"]["A2"].value == "a"

    review = wb["Needs Review"]
    id_col = _header_col(review, "receipt_id")
    # r3's explicit id wins over the metadata's own receipt_id.
    assert review.cell(row=2, column=id_col).value == "c"


# --------------------------------------------------------------------------- #
# The buyer, and the rows nobody bought.
#
# A BIR sales invoice names who it was sold TO -- distinct from the merchant who
# sold it -- and prints product rows that stay blank when nothing on that line
# was bought. Both reach the workbook, but not in the same way: the buyer gets
# its own columns, while a blank pre-printed row is deliberately kept OUT of the
# LineItems sheet. It was transcribed so nothing on the paper is lost; it is not
# a purchase, and a ledger that lists one is wrong about what was bought.
# --------------------------------------------------------------------------- #


def _purchase(description: str, amount: Decimal, *, position: int = 1) -> LineItem:
    return LineItem(
        position=position,
        description_raw=description,
        qty=Decimal("1"),
        unit_price=amount,
        line_total=amount,
    )


def _template(description: str, *, position: int = 1) -> LineItem:
    """A pre-printed product row left blank on the form: transcribed, not bought."""
    return LineItem(position=position, description_raw=description, is_template_row=True)


def _bir_receipt(
    *,
    buyer_name: str | None = None,
    buyer_tax_id: str | None = None,
    line_items: list[LineItem] | None = None,
    total: Decimal = Decimal("2000.00"),
) -> ReceiptExtraction:
    return ReceiptExtraction(
        merchant=Merchant(name="METRO OIL SUBIC, INC."),
        buyer=Buyer(name=buyer_name, tax_id=buyer_tax_id),
        receipt=ReceiptMeta(date="2026-03-23", currency="PHP"),
        line_items=(
            [_purchase("DieselPlus", Decimal("2000.00"))] if line_items is None else line_items
        ),
        totals=Totals(total=total),
        payment=Payment(method="cash"),
    )


def _book(tmp_path, receipts, rows=None):
    out = tmp_path / "bir.xlsx"
    export_workbook(receipts, out_path=out, rows=rows)
    return openpyxl.load_workbook(out)


def _column_values(ws, name: str) -> list:
    """Every data cell under the named header, top to bottom."""
    col = _header_col(ws, name)
    return [ws.cell(row=row, column=col).value for row in range(2, ws.max_row + 1)]


def test_the_review_sheet_carries_the_buyer(tmp_path):
    wb = _book(
        tmp_path,
        [_bir_receipt(buyer_name="IDEAL SOURCE")],
        rows=[
            ReceiptExportRow(
                receipt_id="r1",
                status=ReceiptStatus.NEEDS_REVIEW,
                review_reason="verify the buyer",
                review_priority=1,
            )
        ],
    )
    assert _column_values(wb["Needs Review"], "buyer") == ["IDEAL SOURCE"]


def test_the_receipts_sheet_carries_the_buyer_name_and_tin(tmp_path):
    ws = _book(
        tmp_path,
        [_bir_receipt(buyer_name="IDEAL SOURCE", buyer_tax_id="008-123-456-000")],
    )["Receipts"]

    assert _column_values(ws, "buyer") == ["IDEAL SOURCE"]
    assert _column_values(ws, "buyer_tax_id") == ["008-123-456-000"]
    # A TIN is digits and separators, never a quantity: text format, or Excel
    # coerces it and eats any leading zero (§13.5, same as card_last4).
    assert ws.cell(row=2, column=_header_col(ws, "buyer_tax_id")).number_format == "@"


def test_a_receipt_that_names_no_buyer_leaves_the_buyer_cells_empty(tmp_path):
    """A receipt that names no buyer gets blank cells, never the text "None".

    Blankness is all this can pin, and all any test of a saved workbook can:
    openpyxl drops an empty-string cell on write, so "" and None reload
    identically. Null-versus-empty-string is a real distinction on the internal
    path -- ``receipts.buyer_name_raw`` stores NULL, never "" -- but it is
    invisible at this boundary by construction, so no assertion here can hold
    the exporter to it.
    """
    ws = _book(tmp_path, [_bir_receipt()])["Receipts"]

    assert _column_values(ws, "buyer") == [None]
    assert _column_values(ws, "buyer_tax_id") == [None]


def test_a_template_row_is_not_exported_as_a_purchase(tmp_path):
    """An accounting ledger listing something nobody bought is a defect."""
    wb = _book(
        tmp_path,
        [
            _bir_receipt(
                line_items=[
                    _template("MaxiPower", position=1),
                    _purchase("DieselPlus", Decimal("2000.00"), position=2),
                ]
            )
        ],
    )
    assert _column_values(wb["LineItems"], "description") == ["DieselPlus"]


def test_the_items_count_counts_purchases_not_pre_printed_rows(tmp_path):
    """`items` must agree with the LineItems sheet, or the workbook contradicts itself."""
    wb = _book(
        tmp_path,
        [
            _bir_receipt(
                line_items=[
                    _template("MaxiPower", position=1),
                    _purchase("DieselPlus", Decimal("2000.00"), position=2),
                ]
            )
        ],
    )
    assert _column_values(wb["Receipts"], "items") == [1]
    assert wb["LineItems"].max_row == 2  # header + the one purchase


def test_a_receipt_of_nothing_but_template_rows_still_exports_the_receipt(tmp_path):
    """A BIR form with nothing filled in is a real scan.

    The receipt is a row on Receipts -- it was scanned, it exists, and dropping
    it would hide a receipt somebody has to look at. Its pre-printed rows are
    not purchases, so the ledger gets none of them and the count says zero.
    """
    wb = _book(
        tmp_path,
        [
            _bir_receipt(
                total=Decimal("0.00"),
                line_items=[
                    _template("MaxiPower", position=1),
                    _template("DieselPlus", position=2),
                ],
            )
        ],
    )
    assert _column_values(wb["Receipts"], "merchant") == ["METRO OIL SUBIC, INC."]
    assert _column_values(wb["Receipts"], "items") == [0]
    assert wb["LineItems"].max_row == 1  # header only: no purchases to list
    assert _column_values(wb["LineItems"], "description") == []
