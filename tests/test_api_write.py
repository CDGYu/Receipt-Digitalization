"""Tests for the review API's write routes (P4.T5, spec §14.9).

``pytest.importorskip("fastapi")``/``pytest.importorskip("openpyxl")`` keep
the base test suite offline, matching ``tests/test_api_read.py`` and
``tests/test_xlsx.py``.

Everything runs against a file-backed SQLite database, a ``LocalStorage``
rooted at ``tmp_path``, and a fake ``submit`` that appends to a list instead
of touching Redis/RQ -- the same pattern ``test_api_read.py`` uses.

Fixture design note: ``reviewer_client``/``session_factory`` seed **no
receipts at all** -- only the two accounts. That is load-bearing for
``test_a_rejected_upload_writes_no_row_and_queues_nothing``, which asserts
the *entire* ``receipts`` table is empty after a rejected upload. Every test
that needs a receipt asks for it explicitly (``receipt_id``,
``other_receipt_id``, ``pending_receipt_id``, ``task_id``), and each of
those fixtures inserts its own row into the same shared database the moment
it is requested -- so a test that does not name one never sees it.
``admin_client`` is the one fixture that seeds by side effect (via
``receipt_id``/``other_receipt_id``, plus an open review task): every route
that needs an admin also needs *something* for the admin to act on --
two exportable receipts (so the row-cap test can force an overflow with
``_EXPORT_MAX_ROWS = 1``) and one open task (so ``/review/next`` has
something to claim). ``reviewer_client`` never triggers this, so the upload
tests stay receipt-free.
"""

from __future__ import annotations

import glob
import io
import logging
import os
import tempfile
import uuid
from datetime import date, time
from decimal import Decimal

import pytest

pytest.importorskip("fastapi")
pytest.importorskip("openpyxl")

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import event, select  # noqa: E402

import receipts.review.api as api_module  # noqa: E402
from config.settings import Settings  # noqa: E402
from receipts.ingest.storage import LocalStorage, make_image_key  # noqa: E402
from receipts.persist.models import (  # noqa: E402
    Base,
    Correction,
    Merchant,
    Receipt,
    ReviewState,
    ReviewTask,
)
from receipts.persist.session import make_engine, make_session_factory  # noqa: E402
from receipts.persist.users import ROLE_ADMIN, ROLE_REVIEWER, create_user  # noqa: E402
from receipts.review.api import create_app  # noqa: E402
from receipts.review.queue import enqueue_review, next_task  # noqa: E402
from receipts.score.confidence import ReceiptStatus  # noqa: E402

#: Minimal but genuinely JPEG-sniffable bytes (see
#: ``receipts.ingest.ingest._sniff_content_type``): the ``\xff\xd8`` header
#: is all the validator inspects.
JPEG_BYTES = b"\xff\xd8\xff\xe0" + b"\x00" * 128 + b"\xff\xd9"

MAIN_RECEIPT = uuid.uuid4()  # receipt_id -- needs_review, image = JPEG_BYTES
OTHER_RECEIPT = uuid.uuid4()  # other_receipt_id -- auto_approved
PENDING_RECEIPT = uuid.uuid4()  # pending_receipt_id -- pending


# --------------------------------------------------------------------------- #
# Fixtures
# --------------------------------------------------------------------------- #


@pytest.fixture()
def storage(tmp_path) -> LocalStorage:
    return LocalStorage(tmp_path / "blobs")


@pytest.fixture()
def session_factory(tmp_path):
    """A fresh database with the two accounts and **no receipts**."""
    engine = make_engine(f"sqlite:///{(tmp_path / 'receipts.db').as_posix()}")
    Base.metadata.create_all(engine)
    factory = make_session_factory(engine)
    with factory() as session:
        create_user(session, "alice", "pw-alice", ROLE_REVIEWER)
        create_user(session, "bob", "pw-bob", ROLE_ADMIN)
        session.commit()
    return factory


@pytest.fixture()
def settings() -> Settings:
    """Hermetic settings: a developer's ``.env`` must not steer these tests."""
    return Settings(
        _env_file=None,
        session_secret="test-secret",
        session_cookie_secure=False,
        receipts_api_key="s3cret-machine-key",
    )


@pytest.fixture()
def submitted() -> list:
    """What a fake ``submit`` records instead of touching Redis/RQ."""
    return []


@pytest.fixture()
def app(session_factory, storage, settings, submitted):
    return create_app(
        session_factory=session_factory,
        storage=storage,
        submit=submitted.append,
        settings=settings,
    )


def _logged_in(app, username: str, password: str) -> TestClient:
    client = TestClient(app)
    response = client.post("/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200
    return client


@pytest.fixture()
def reviewer_client(app) -> TestClient:
    return _logged_in(app, "alice", "pw-alice")


@pytest.fixture()
def key_client(app) -> TestClient:
    """No session -- the caller supplies ``X-API-Key`` itself, per call."""
    return TestClient(app)


@pytest.fixture()
def receipt_id(session_factory, storage) -> uuid.UUID:
    """A ``needs_review`` receipt whose image bytes are actually in storage."""
    key = make_image_key(MAIN_RECEIPT, "original")
    with session_factory() as session:
        session.add(
            Receipt(
                id=MAIN_RECEIPT,
                status=ReceiptStatus.NEEDS_REVIEW,
                confidence=Decimal("0.700"),
                merchant_name_raw="COFFEE CO",
                txn_date=date(2026, 7, 1),
                # Seconds on purpose: they are what tells an ``isoformat()``
                # round-trip apart from a lossy ``%H:%M`` one.
                txn_time=time(14, 30, 45),
                currency="USD",
                total=Decimal("12.50"),
                image_key=key,
                image_phash="",
            )
        )
        session.commit()
    storage.put(key, JPEG_BYTES, "image/jpeg")
    return MAIN_RECEIPT


@pytest.fixture()
def other_receipt_id(session_factory) -> uuid.UUID:
    with session_factory() as session:
        session.add(
            Receipt(
                id=OTHER_RECEIPT,
                status=ReceiptStatus.AUTO_APPROVED,
                confidence=Decimal("0.930"),
                merchant_name_raw="TOTAL WINE",
                txn_date=date(2026, 7, 2),
                currency="USD",
                total=Decimal("40.00"),
                image_key=make_image_key(OTHER_RECEIPT, "original"),
                image_phash="",
            )
        )
        session.commit()
    return OTHER_RECEIPT


@pytest.fixture()
def pending_receipt_id(session_factory) -> uuid.UUID:
    with session_factory() as session:
        session.add(
            Receipt(
                id=PENDING_RECEIPT,
                status=ReceiptStatus.PENDING,
                confidence=Decimal("0"),
                image_key=make_image_key(PENDING_RECEIPT, "original"),
                image_phash="",
            )
        )
        session.commit()
    return PENDING_RECEIPT


@pytest.fixture()
def task_id(session_factory, receipt_id) -> uuid.UUID:
    """The seeded receipt's review task, already claimed by alice -- as if
    she had already called ``GET /review/next`` before these tests exercise
    ``POST /review/{id}/complete``.
    """
    with session_factory() as session:
        enqueue_review(session, receipt_id, reason="quick verify", priority=2)
        task = next_task(session, assignee="alice")
        session.commit()
        return task.id


@pytest.fixture()
def admin_client(app, session_factory, receipt_id, other_receipt_id) -> TestClient:
    """Logged in as bob, with something for an admin to act on: two
    exportable receipts and one open review task.
    """
    with session_factory() as session:
        enqueue_review(session, receipt_id, reason="quick verify", priority=2)
        session.commit()
    return _logged_in(app, "bob", "pw-bob")


@pytest.fixture()
def client_max_1mb(tmp_path) -> TestClient:
    """A reviewer-authenticated client whose app enforces a 1 MB upload cap."""
    settings = Settings(
        _env_file=None,
        session_secret="test-secret",
        session_cookie_secure=False,
        max_upload_mb=1,
    )
    engine = make_engine(f"sqlite:///{(tmp_path / 'small.db').as_posix()}")
    Base.metadata.create_all(engine)
    factory = make_session_factory(engine)
    with factory() as session:
        create_user(session, "alice", "pw-alice", ROLE_REVIEWER)
        session.commit()
    small_app = create_app(
        session_factory=factory,
        storage=LocalStorage(tmp_path / "small-blobs"),
        submit=lambda job: None,
        settings=settings,
    )
    return _logged_in(small_app, "alice", "pw-alice")


@pytest.fixture()
def empty_reviewer_client(tmp_path) -> TestClient:
    """A second, genuinely empty database -- no receipts, no queue tasks."""
    settings = Settings(_env_file=None, session_secret="test-secret", session_cookie_secure=False)
    engine = make_engine(f"sqlite:///{(tmp_path / 'empty.db').as_posix()}")
    Base.metadata.create_all(engine)
    factory = make_session_factory(engine)
    with factory() as session:
        create_user(session, "alice", "pw-alice", ROLE_REVIEWER)
        session.commit()
    empty_app = create_app(
        session_factory=factory,
        storage=LocalStorage(tmp_path / "empty-blobs"),
        submit=lambda job: None,
        settings=settings,
    )
    return _logged_in(empty_app, "alice", "pw-alice")


# --------------------------------------------------------------------------- #
# Brief step 1, verbatim
# --------------------------------------------------------------------------- #


def test_upload_writes_a_pending_row_then_queues(reviewer_client, session_factory, submitted):
    response = reviewer_client.post("/upload", files={"file": ("r.jpg", JPEG_BYTES, "image/jpeg")})
    assert response.status_code == 202
    receipt_id = uuid.UUID(response.json()["receipt_id"])
    with session_factory() as session:
        row = session.get(Receipt, receipt_id)
    # The row exists BEFORE the worker runs: a job the queue loses is visible
    # as a stuck pending row, not a blob with nothing in the database.
    assert row.status is ReceiptStatus.PENDING
    assert [job.id for job in submitted] == [receipt_id]


def test_a_rejected_upload_writes_no_row_and_queues_nothing(
    reviewer_client, session_factory, submitted
):
    response = reviewer_client.post(
        "/upload", files={"file": ("notes.txt", b"hello", "text/plain")}
    )
    assert response.status_code == 400
    with session_factory() as session:
        assert session.query(Receipt).count() == 0
    assert submitted == []


def test_upload_honours_the_configured_size_limit(client_max_1mb):
    big = b"\xff\xd8" + b"\x00" * (2 * 1024 * 1024)
    response = client_max_1mb.post("/upload", files={"file": ("r.jpg", big, "image/jpeg")})
    assert response.status_code == 400


def test_the_api_key_can_upload_but_nothing_else(key_client, receipt_id):
    headers = {"X-API-Key": "s3cret-machine-key"}
    assert key_client.post("/upload", files={"file": ("r.jpg", JPEG_BYTES, "image/jpeg")},
                           headers=headers).status_code == 202
    assert key_client.patch(f"/receipts/{receipt_id}", json={"totals": {"total": "1.00"}},
                            headers=headers).status_code == 401


def test_patch_writes_a_correction_attributed_to_the_session_user(
    reviewer_client, session_factory, receipt_id
):
    response = reviewer_client.patch(
        f"/receipts/{receipt_id}", json={"totals": {"total": "1234.56"}}
    )
    assert response.status_code == 200
    with session_factory() as session:
        correction = session.scalars(select(Correction)).one()
    # The entire reason session auth was chosen over a shared key.
    assert correction.corrected_by == "alice"
    assert correction.value_after == "1234.56"


def test_a_buyer_correction_survives_the_route_and_comes_back_redacted(
    reviewer_client, session_factory, receipt_id
):
    """The buyer is correctable *through the API*, not only in the repository.

    ``CorrectionPatch`` names ``merchant``/``receipt``/``totals``/``payment``/
    ``meta``/``line_items`` and does **not** name ``buyer``; it reaches
    ``apply_corrections`` only because every level of that model is
    ``extra="allow"`` and the route dumps with ``exclude_unset=True``. That is
    a deliberate design (one error currency for "unknown field"), but it is
    also exactly the kind of pass-through that a later tightening of the patch
    model would silently break -- with the repository tests still green,
    because they never cross this layer. This is that binding.

    The PAN is here for the same reason the ``payment_method`` test above
    carries one: ``buyer_name_raw`` is a new free-text column served over the
    API, and a reviewer retyping a "Sold To" block off a slip can put a card
    number in it. Redaction is bound one layer down in
    ``tests/test_repository.py``; this checks the route agrees.
    """
    response = reviewer_client.patch(
        f"/receipts/{receipt_id}",
        json={"buyer": {"name": "IDEAL SOURCE 4111111111111111", "tax_id": "123-456-789-000"}},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["buyer"]["tax_id"] == "123-456-789-000"
    assert "4111111111111111" not in body["buyer"]["name"]
    assert body["buyer"]["name"].startswith("IDEAL SOURCE ")

    with session_factory() as session:
        stored = session.get(Receipt, receipt_id)
        assert stored.buyer_tax_id == "123-456-789-000"
        assert "4111111111111111" not in stored.buyer_name_raw
        # The audit copy is the one nothing later scrubs.
        rows = session.scalars(
            select(Correction).where(Correction.field_path == "buyer.name")
        ).all()
        assert len(rows) == 1
        assert "4111111111111111" not in rows[0].value_after


def test_the_time_the_detail_returns_patches_back_unchanged(
    reviewer_client, session_factory, receipt_id
):
    """``receipt.time`` is correctable, so the rendering has to round-trip.

    A reviewer's screen is populated from ``GET /receipts/{id}`` and its edits
    go back through ``PATCH``. If the two disagree about how a ``time`` is
    written, a reviewer who merely *confirms* an untouched receipt rewrites the
    stored value -- and ``apply_corrections`` would log that as a correction
    they never made. Asserting the ``corrections`` table stayed empty is what
    makes this test see a lossy rendering: ``14:30:45`` rendered as ``14:30``
    would patch back as a genuine change and write a row.
    """
    rendered = reviewer_client.get(f"/receipts/{receipt_id}").json()["txn_time"]

    response = reviewer_client.patch(
        f"/receipts/{receipt_id}", json={"receipt": {"time": rendered}}
    )

    assert response.status_code == 200
    assert response.json()["txn_time"] == rendered
    with session_factory() as session:
        assert session.get(Receipt, receipt_id).txn_time == time(14, 30, 45)
        assert session.scalars(select(Correction)).all() == []


def test_the_newly_exposed_payment_method_never_returns_a_full_pan(
    reviewer_client, receipt_id
):
    """§18 at the layer that now serves the column (P5.T3b, fix round 1).

    ``payment_method`` became readable over the API in this task, and its
    redaction was bound only one layer down, in ``tests/test_repository.py``.
    Measured: with ``redact_pan`` neutered to the identity function, all 116
    tests across ``test_api_read.py``/``test_api_write.py``/
    ``test_review_queue.py`` still passed while ``test_repository.py`` failed
    11 -- so nothing at the route layer would have noticed a third writer of
    this column, or a serializer that reached around the repository. This is
    that binding.

    ``payment.method`` is the motivating path named in ``_plan_change``'s own
    docstring: it is where a reviewer types "VISA 4111111111111111" straight
    off the slip.
    """
    response = reviewer_client.patch(
        f"/receipts/{receipt_id}", json={"payment": {"method": "VISA 4111111111111111"}}
    )

    assert response.status_code == 200
    assert response.json()["payment_method"] == "VISA ************1111"
    # ...and it stays masked on the way back out, not just in the echo.
    body = reviewer_client.get(f"/receipts/{receipt_id}")
    assert body.json()["payment_method"] == "VISA ************1111"
    assert "4111111111111111" not in body.text


@pytest.mark.parametrize(
    "typed",
    [
        "VISA 4111.1111.1111.1111",
        "VISA 4111_1111_1111_1111",
        "VISA 4111/1111/1111/1111",
        "VISA 4111,1111,1111,1111",
        "VISA 4111 1111-1111.1111",
    ],
)
def test_a_dotted_pan_is_masked_in_the_row_the_body_and_the_audit_copy(
    reviewer_client, session_factory, receipt_id, typed
):
    """§18 end to end, through the route a reviewer actually types into.

    Three places, because fixing one of them later cannot fix the others: the
    ``receipts`` row, the ``GET`` body, and ``corrections.value_after``. The
    audit table is durable and append-only -- a correction row is never
    rewritten -- so a PAN that lands there is not recoverable from by scrubbing
    the receipt afterwards. ``_plan_change``'s own docstring calls it "precisely
    the copy nothing later scrubs".

    Only spaces and hyphens were separators until this test existed, so every
    parametrisation below stored the whole card number verbatim in all three.
    """
    response = reviewer_client.patch(
        f"/receipts/{receipt_id}", json={"payment": {"method": typed}}
    )
    assert response.status_code == 200

    digits = "".join(ch for ch in typed if ch.isdigit())
    assert response.json()["payment_method"] == "VISA ************1111"

    body = reviewer_client.get(f"/receipts/{receipt_id}")
    assert body.json()["payment_method"] == "VISA ************1111"
    assert digits not in body.text
    assert typed not in body.text

    with session_factory() as session:
        assert session.get(Receipt, receipt_id).payment_method == "VISA ************1111"
        correction = session.scalars(select(Correction)).one()
        assert correction.value_after == "VISA ************1111"
        assert digits not in correction.value_after
        assert typed not in correction.value_after


def test_patch_rejects_a_json_float_for_money(reviewer_client, receipt_id):
    response = reviewer_client.patch(f"/receipts/{receipt_id}", json={"totals": {"total": 1234.56}})
    assert response.status_code == 422
    assert "string" in response.text


def test_patch_with_an_unmappable_path_changes_nothing(
    reviewer_client, session_factory, receipt_id
):
    response = reviewer_client.patch(
        f"/receipts/{receipt_id}", json={"nonsense": {"field": "x"}}
    )
    assert response.status_code == 400
    with session_factory() as session:
        assert session.query(Correction).count() == 0


# --------------------------------------------------------------------------- #
# The correction is re-checked, on both routes that serve the detail (Task 4)
# --------------------------------------------------------------------------- #


@pytest.mark.parametrize(
    ("configured", "patch_body"),
    [
        pytest.param(
            {"expected_buyer_name": "IDEAL SOURCE"},
            {"buyer": {"name": "SOMEONE ELSE"}},
            id="expected_buyer_name",
        ),
        pytest.param(
            {"expected_buyer_tax_id": "123-456-789-000"},
            {"buyer": {"tax_id": "999-999-999-999"}},
            id="expected_buyer_tax_id",
        ),
    ],
)
def test_a_correction_is_re_checked_by_both_routes_that_serve_the_detail(
    session_factory, storage, settings, submitted, receipt_id, configured, patch_body
):
    """The reviewer edits the buyer, and the panel answers about the edit.

    Two things at once, because they are the same fact seen twice. Nothing is
    stored: the ``PATCH`` response and the ``GET`` that follows it both compute
    ``current_findings`` from the row, so both must show ``R015`` and neither
    can be reading a copy written at extraction time -- ``MAIN_RECEIPT`` was
    never extracted and has no findings row at all.

    **Exactly one ``expected_buyer_*`` field is configured per case, and that is
    the point.** ``expects_a_buyer`` (``validate/rules.py``) gates R014/R015 on
    *either* being non-blank, so a case that set both would still fire with one
    of the two kwargs dropped from a ``receipt_detail`` call site, and the loss
    would be invisible. One at a time makes each kwarg, at each of the two call
    sites, individually load-bearing.

    The ``before`` assertion is what keeps this from passing vacuously: R015 is
    silent on the untouched receipt (no buyer name and no buyer TIN is "not
    read", which is R014's subject, not a mismatch), so its appearance is caused
    by the edit rather than by the fixture.
    """
    client = _logged_in(
        create_app(
            session_factory=session_factory,
            storage=storage,
            submit=submitted.append,
            settings=settings.model_copy(update=configured),
        ),
        "alice",
        "pw-alice",
    )

    before = client.get(f"/receipts/{receipt_id}").json()
    assert before["findings"] == []
    assert "R015" not in [f["rule_id"] for f in before["current_findings"]]

    patched = client.patch(f"/receipts/{receipt_id}", json=patch_body)
    assert patched.status_code == 200
    assert "R015" in [f["rule_id"] for f in patched.json()["current_findings"]]

    after = client.get(f"/receipts/{receipt_id}").json()
    assert "R015" in [f["rule_id"] for f in after["current_findings"]]
    # History is history: re-validation writes nothing, here or anywhere.
    assert after["findings"] == []


def test_image_url_is_signed_and_the_blob_streams(reviewer_client, receipt_id):
    url = reviewer_client.get(f"/receipts/{receipt_id}/image").json()["url"]
    assert reviewer_client.get(url).content == JPEG_BYTES


def test_a_valid_signature_for_a_missing_blob_is_404_not_500(reviewer_client, other_receipt_id):
    """``other_receipt_id`` has a real ``receipts`` row and a real
    ``image_key``, but its fixture never writes bytes to storage -- the
    same shape a receipt whose blob was evicted or never uploaded would
    have. The signature is genuinely valid (right receipt, right variant,
    right secret); only the file is absent. Before fix round 1 (F5) this
    reached ``storage.get`` unguarded and ``LocalStorage.get`` raised
    ``FileNotFoundError`` straight through the route as an unhandled 500 --
    on the one route in the service that requires no authentication at
    all.
    """
    url = reviewer_client.get(f"/receipts/{other_receipt_id}/image").json()["url"]
    assert reviewer_client.get(url).status_code == 404


def test_a_tampered_signature_is_rejected(reviewer_client, receipt_id, other_receipt_id):
    """Two tampering shapes, not expiry: re-pointing the URL at a different
    receipt (the signature covers ``receipt_id``, so it cannot be replayed
    against one it was not minted for), and corrupting ``exp`` itself (the
    replacement prefixes a digit onto the existing value, producing a
    *larger*, still-unexpired ``exp`` -- but paired with a signature that no
    longer matches it). Neither case waits for real expiry; see
    `test_a_correctly_signed_link_is_refused_once_it_expires` (fix round 1,
    F7) for that -- this test used to be misnamed as covering it.
    """
    url = reviewer_client.get(f"/receipts/{receipt_id}/image").json()["url"]
    swapped_id = url.replace(str(receipt_id), str(other_receipt_id))
    assert reviewer_client.get(swapped_id).status_code == 403
    assert reviewer_client.get(url.replace("exp=", "exp=1")).status_code == 403


def test_a_correctly_signed_link_is_refused_once_it_expires(session_factory, storage, receipt_id):
    """A signature that is valid in every other respect -- right receipt,
    right variant, right secret -- is still refused once ``exp`` has
    passed. This is the genuine expiry case
    `test_a_tampered_signature_is_rejected` was misnamed for (fix round 1,
    F7): a separate app, sharing the same database and storage, whose
    ``IMAGE_URL_TTL_S`` is negative -- so the link is already expired at
    the instant ``sign_url`` mints it, with no need to sleep in the test.
    """
    expiring_settings = Settings(
        _env_file=None,
        session_secret="test-secret",
        session_cookie_secure=False,
        image_url_ttl_s=-5,
    )
    expiring_app = create_app(
        session_factory=session_factory,
        storage=storage,
        submit=lambda job: None,
        settings=expiring_settings,
    )
    client = _logged_in(expiring_app, "alice", "pw-alice")
    url = client.get(f"/receipts/{receipt_id}/image").json()["url"]
    assert client.get(url).status_code == 403


def test_review_next_claims_one_task_per_caller(reviewer_client, admin_client):
    first = reviewer_client.get("/review/next").json()["task"]
    second = admin_client.get("/review/next").json()["task"]
    assert first["id"] != (second or {}).get("id")


def test_review_next_returns_the_same_task_when_a_reviewer_reloads(
    reviewer_client, session_factory, receipt_id, other_receipt_id
):
    """The page-reload case ADR-0016 exists for, end to end.

    Nothing the *reviewer* can reach releases a claim: ``_claim_stmt`` selects
    ``state == OPEN`` only, ``enqueue_review`` reopens a task only when it is
    ``DONE``, and ``POST /review/{id}/complete`` closes rather than releases.
    ``POST /review/{id}/release`` does release one, but it is admin-only
    (ADR-0025) and is not on this path. So without resume, a reviewer who
    reloaded claimed a *second* task and stranded the first until an admin
    intervened. Two tasks are queued here rather than one precisely so a
    regression could take the second: with one task the buggy behaviour and
    the fixed one both return the same id.
    """
    with session_factory() as session:
        enqueue_review(session, receipt_id, reason="quick verify", priority=2)
        enqueue_review(session, other_receipt_id, reason="urgent: no total", priority=0)
        session.commit()

    first = reviewer_client.get("/review/next").json()
    second = reviewer_client.get("/review/next").json()

    assert first["task"]["id"] == second["task"]["id"]
    assert first["receipt"]["id"] == second["receipt"]["id"]
    assert second["task"]["assigned_to"] == "alice"
    with session_factory() as session:
        held = session.scalars(
            select(ReviewTask).where(ReviewTask.state == ReviewState.IN_PROGRESS)
        ).all()
        assert [str(task.id) for task in held] == [first["task"]["id"]]


def test_review_next_on_an_empty_queue_returns_null(empty_reviewer_client):
    assert empty_reviewer_client.get("/review/next").json()["task"] is None


def test_a_reviewer_cannot_complete_someone_elses_task(reviewer_client, session_factory, task_id):
    with session_factory() as session:
        task = session.get(ReviewTask, task_id)
        task.assigned_to = "bob"
        session.commit()
    assert reviewer_client.post(f"/review/{task_id}/complete").status_code == 403


def test_an_admin_can_complete_a_task_assigned_to_someone_else(
    admin_client, session_factory, task_id
):
    with session_factory() as session:
        task = session.get(ReviewTask, task_id)
        task.assigned_to = "alice"
        session.commit()
    assert admin_client.post(f"/review/{task_id}/complete").status_code == 200


def test_completing_an_unknown_task_is_404(reviewer_client):
    assert reviewer_client.post(f"/review/{uuid.uuid4()}/complete").status_code == 404


def test_double_complete_does_not_move_closed_at(reviewer_client, session_factory, task_id):
    reviewer_client.post(f"/review/{task_id}/complete")
    with session_factory() as session:
        first_closed = session.get(ReviewTask, task_id).closed_at
    reviewer_client.post(f"/review/{task_id}/complete")
    with session_factory() as session:
        assert session.get(ReviewTask, task_id).closed_at == first_closed


def test_review_complete_requires_authentication(app, task_id):
    """The one non-``/health``, non-blob route that had no pin for "no
    credentials at all" (fix round 1, F4). Its real permission rule --
    assignee or admin -- cannot be expressed as a row in
    ``tests/test_api_read.py``'s role/actor matrix (a reviewer holding an
    allowed role can still get 403 there), which is why it was never added
    as one; that is not a reason for the baseline "every non-``/health``
    route needs 401 without credentials" property to go unpinned. Neither
    an anonymous caller nor the machine upload key -- which authorizes
    ``POST /upload`` and nothing else -- may complete a task.
    """
    anonymous = TestClient(app)
    assert anonymous.post(f"/review/{task_id}/complete").status_code == 401

    keyed = TestClient(app)
    keyed.headers.update({"X-API-Key": "s3cret-machine-key"})
    assert keyed.post(f"/review/{task_id}/complete").status_code == 401


# --------------------------------------------------------------------------- #
# ADR-0025: `POST /review/{task_id}/release`, admin only.
#
# Deliberately NOT a row in `tests/test_api_read.py`'s READ_ROUTES matrix, and
# for a different reason than the `/complete` exclusion that file already
# documents. `test_auth_matrix` builds every URL with
# `path.format(id=receipt_id)` and asserts 200 for an allowed actor, but this
# route's path parameter is a **review task** id, not a receipt id -- a receipt
# id substituted there is a legitimate 404, so an admin row would fail on a
# correct implementation. The matrix's shape cannot express this route at all;
# its auth is pinned behaviourally here instead
# (`test_a_reviewer_cannot_release_a_task`, `test_release_requires_authentication`).
# --------------------------------------------------------------------------- #


def test_an_admin_can_release_a_claimed_task(admin_client, session_factory, task_id):
    response = admin_client.post(f"/review/{task_id}/release")

    assert response.status_code == 200
    body = response.json()
    assert body["state"] == "open"
    assert body["assigned_to"] is None
    # A sibling key, not a replacement: assigned_to says who holds it (nobody),
    # released_from says who held it.
    assert body["released_from"] == "alice"

    with session_factory() as session:
        stored = session.get(ReviewTask, task_id)
        assert stored.state.value == "open"
        assert stored.assigned_to is None


def test_a_reviewer_cannot_release_a_task(reviewer_client, task_id):
    response = reviewer_client.post(f"/review/{task_id}/release")

    assert response.status_code == 403
    assert response.json()["error"]["message"] == "insufficient role"


def test_release_requires_authentication(app, task_id):
    """Both credential-less callers, matching
    ``test_review_complete_requires_authentication``: an anonymous one and the
    machine ``X-API-Key`` client, which authorizes ``POST /upload`` and nothing
    else (fix wave, F11). The key row is pinned here because this route cannot
    join ``test_api_read.py``'s ``test_auth_matrix`` -- see the block comment
    above for why -- which is the same reason ``POST /review/{id}/complete``'s
    key row is pinned by hand a few tests up, in this file. Routes whose rule
    *is* a role/actor predicate get theirs from one of that module's matrices
    instead.
    """
    anonymous = TestClient(app).post(f"/review/{task_id}/release")

    assert anonymous.status_code == 401
    assert anonymous.json()["error"]["message"] == "authentication required"

    keyed_client = TestClient(app)
    keyed_client.headers.update({"X-API-Key": "s3cret-machine-key"})
    keyed = keyed_client.post(f"/review/{task_id}/release")

    assert keyed.status_code == 401
    assert keyed.json()["error"]["message"] == "authentication required"


def test_releasing_an_unknown_task_is_404(admin_client):
    """Not 400. ``release_task``'s own ValueError for an unknown id would render
    400 through the error handler, so the route keeps its own existence check --
    "no such task" and "that task cannot be released" are different answers.
    """
    missing = uuid.uuid4()

    response = admin_client.post(f"/review/{missing}/release")

    assert response.status_code == 404
    assert str(missing) in response.json()["error"]["message"]


def test_releasing_a_closed_task_is_400(admin_client, task_id):
    assert admin_client.post(f"/review/{task_id}/complete").status_code == 200

    response = admin_client.post(f"/review/{task_id}/release")

    assert response.status_code == 400
    assert str(task_id) in response.json()["error"]["message"]


def test_a_released_reviewer_gets_403_on_complete(admin_client, reviewer_client, task_id):
    """The milestone's headline claim: this is the exact 403 ADR-0024's terminal
    `taken` state was built for, produced by a real release rather than a
    hand-set fixture. Until this route shipped, that UI path was reachable only
    by tests that set ``assigned_to`` by hand.
    """
    assert admin_client.post(f"/review/{task_id}/release").status_code == 200

    response = reviewer_client.post(f"/review/{task_id}/complete")

    assert response.status_code == 403
    assert (
        response.json()["error"]["message"]
        == "only the assignee or an admin may complete this task"
    )


def test_the_release_is_logged_without_the_tasks_reason(
    admin_client, session_factory, task_id, caplog
):
    """``reason`` is built from exception text and is redacted only at
    ``enqueue_review``'s sink, so putting it in a log line would extend
    ADR-0022's egress inventory. Ids and usernames only.
    """
    with session_factory() as session:
        task = session.get(ReviewTask, task_id)
        task.reason = "SENTINEL-REASON-TEXT"
        session.commit()

    with caplog.at_level(logging.INFO, logger="receipts.review.api"):
        assert admin_client.post(f"/review/{task_id}/release").status_code == 200

    lines = [r.getMessage() for r in caplog.records if "released" in r.getMessage()]
    assert len(lines) == 1
    assert str(task_id) in lines[0]
    assert "alice" in lines[0]
    assert "bob" in lines[0]
    assert "SENTINEL-REASON-TEXT" not in lines[0]


def test_releasing_an_already_open_task_is_not_logged_as_a_release(
    admin_client, session_factory, receipt_id, caplog
):
    """A no-op must not be announced as a release (fix wave, F9).

    The comment at the log site is careful that a *rolled-back* release is never
    announced as one; a no-op was, as ``released from None``. ADR-0025 §3 calls
    this line the only durable trace of a release, so a false positive in it is
    an audit defect: nothing queryable records who released what, and the
    response body that does distinguish the two is not kept anywhere.

    ``admin_client`` seeds an **open**, never-claimed task for ``receipt_id``
    (its own fixture does the ``enqueue_review``), which is exactly the
    idempotent path -- this test deliberately does not ask for ``task_id``,
    whose fixture would claim that same row for alice.
    """
    with session_factory() as session:
        open_task_id = (
            session.scalars(select(ReviewTask).where(ReviewTask.receipt_id == receipt_id))
            .one()
            .id
        )
        assert session.get(ReviewTask, open_task_id).state is ReviewState.OPEN

    with caplog.at_level(logging.INFO, logger="receipts.review.api"):
        response = admin_client.post(f"/review/{open_task_id}/release")

    assert response.status_code == 200
    assert response.json()["released_from"] is None

    lines = [r.getMessage() for r in caplog.records if str(open_task_id) in r.getMessage()]
    assert len(lines) == 1
    assert "released from None" not in lines[0]
    assert "nothing released" in lines[0]
    assert "bob" in lines[0]  # the admin who acted is still on the record


def test_a_skipped_receipt_stays_recoverable(
    reviewer_client, session_factory, receipt_id, task_id
) -> None:
    """The three properties the review UI's "Skip this receipt" button spends.

    Skip **completes** the held task, leaving the receipt ``needs_review`` with
    a ``DONE`` task. That is survivable only because of the asymmetry measured
    during the Phase 5 review:

        IN_PROGRESS --enqueue_review--> in_progress   claimable by another: False
        DONE        --enqueue_review--> open          claimable by another: True

    So skip converts the one genuinely unrecoverable queue state into a
    recoverable one. All three properties below are true today and **none would
    go red if they stopped being** -- which is the entire reason this test
    exists.
    """
    assert reviewer_client.post(f"/review/{task_id}/complete").status_code == 200

    # (i) still listed as needing review
    listed = reviewer_client.get("/receipts", params={"status": "needs_review"})
    assert listed.status_code == 200
    assert str(receipt_id) in [item["id"] for item in listed.json()["items"]]

    # (ii) still PATCH-able to `reviewed` -- the route never consults ReviewTask,
    #      so even a patch that changes nothing drives the status.
    patched = reviewer_client.patch(f"/receipts/{receipt_id}", json={})
    assert patched.status_code == 200
    assert patched.json()["status"] == "reviewed"

    # (iii) still re-openable, and re-opened OPEN -- claimable by someone else,
    #       not just by alice.
    with session_factory() as session:
        task = enqueue_review(session, receipt_id, reason="reopened", priority=2)
        session.commit()
        assert task.state is ReviewState.OPEN


def test_export_is_admin_only(reviewer_client, admin_client):
    assert reviewer_client.get("/export/xlsx").status_code == 403
    assert admin_client.get("/export/xlsx").status_code == 200


def test_export_writes_all_four_sheets(admin_client, tmp_path):
    response = admin_client.get("/export/xlsx")
    book = load_workbook(io.BytesIO(response.content))
    assert book.sheetnames == ["Receipts", "LineItems", "Needs Review", "Summary"]


def _receipt_ids_in(response) -> set[str]:
    """The receipt_id column of the Receipts sheet, as strings."""
    sheet = load_workbook(io.BytesIO(response.content))["Receipts"]
    return {str(row[0].value) for row in sheet.iter_rows(min_row=2) if row[0].value}


def test_export_excludes_pending_and_rejected_unless_asked(admin_client, pending_receipt_id):
    default_rows = _receipt_ids_in(admin_client.get("/export/xlsx"))
    # A pending row is an upload in flight, not a transaction.
    assert str(pending_receipt_id) not in default_rows

    asked = _receipt_ids_in(admin_client.get("/export/xlsx", params={"status": "pending"}))
    assert str(pending_receipt_id) in asked


def test_export_refuses_rather_than_truncating(admin_client, monkeypatch):
    monkeypatch.setattr(api_module, "_EXPORT_MAX_ROWS", 1)
    response = admin_client.get("/export/xlsx")
    assert response.status_code == 400
    assert "narrow" in response.text.lower()


def test_export_query_pages_without_repeating_or_skipping(
    session_factory, receipt_id, other_receipt_id
):
    """``offset`` walks the total order ``created_at, id`` establishes.

    Paged one row at a time, the union of the pages is exactly the unpaged
    result and no id appears twice. Asserted over ids rather than over ORM
    identities, because two sessions return different instances for one row.
    """
    from receipts.review.serializers import query_export_receipts

    def page(offset: int) -> list[str]:
        with session_factory() as session:
            rows = query_export_receipts(
                session, status=None, merchant_id=None, date_from=None,
                date_to=None, min_confidence=None, limit=1, offset=offset,
            )
            return [str(row.id) for row in rows]

    with session_factory() as session:
        unpaged = [
            str(row.id)
            for row in query_export_receipts(
                session, status=None, merchant_id=None, date_from=None,
                date_to=None, min_confidence=None, limit=100,
            )
        ]

    # Anti-vacuity: a fixture yielding fewer than two rows would let a broken
    # offset pass, because page 0 alone would equal the unpaged result.
    assert len(unpaged) >= 2, "fixture must produce at least two exportable receipts"

    walked: list[str] = []
    for offset in range(len(unpaged)):
        walked.extend(page(offset))

    assert walked == unpaged
    assert len(set(walked)) == len(walked)
    assert page(len(unpaged)) == []


def test_export_query_pages_a_created_at_tie_without_losing_a_row(session_factory):
    """A paged walk over a tied pair returns every row exactly once.

    Both receipts are seeded with one ``created_at``, which is the case a
    paged walk is likeliest to lose a row on. This pins the walk, not the
    tie-break that makes the order total: SQLite returns tied rows in a
    stable order for repeated identical queries, so these assertions stay
    green even with ``Receipt.id`` dropped from the ``ORDER BY``. That clause
    is pinned separately, by asserting the emitted SQL rather than the row
    order, in ``test_export_query_orders_by_created_at_then_id`` below.
    """
    from datetime import UTC, datetime

    from receipts.persist.models import Receipt
    from receipts.review.serializers import query_export_receipts

    shared = datetime(2026, 7, 4, 12, 0, 0, tzinfo=UTC)
    ids = [uuid.uuid4(), uuid.uuid4()]
    with session_factory() as session:
        for index, receipt_uuid in enumerate(ids):
            session.add(
                Receipt(
                    id=receipt_uuid,
                    status=ReceiptStatus.AUTO_APPROVED,
                    confidence=Decimal("0.900"),
                    merchant_name_raw=f"TIED {index}",
                    currency="USD",
                    total=Decimal("1.00"),
                    image_key=make_image_key(receipt_uuid, "original"),
                    image_phash="",
                    created_at=shared,
                )
            )
        session.commit()

    def page(offset: int) -> list[str]:
        with session_factory() as session:
            return [
                str(row.id)
                for row in query_export_receipts(
                    session, status=None, merchant_id=None, date_from=None,
                    date_to=None, min_confidence=None, limit=1, offset=offset,
                )
            ]

    with session_factory() as session:
        unpaged = [
            str(row.id)
            for row in query_export_receipts(
                session, status=None, merchant_id=None, date_from=None,
                date_to=None, min_confidence=None, limit=100,
            )
        ]

    tied = [str(receipt_uuid) for receipt_uuid in ids]
    walked = [one for offset in range(len(unpaged)) for one in page(offset)]
    assert walked == unpaged
    assert sorted(one for one in walked if one in tied) == sorted(tied)


def test_export_query_orders_by_created_at_then_id(session_factory):
    """``created_at`` leads and ``id`` breaks its ties, asserted on the SQL.

    Paging is only safe over a *total* order, and ``created_at`` alone is not
    one: it is ``server_default=sa.func.now()``, which on SQLite is
    ``CURRENT_TIMESTAMP`` at second resolution, so same-second inserts share
    it. No behavioural test in this module can witness the ``id`` tie-break --
    SQLite orders tied rows stably across repeated identical queries, so a
    paged walk agrees with the unpaged list with or without it. Asserting the
    emitted clause instead does not depend on that behaviour: it fails as soon
    as the tie-break leaves the query, whichever engine compiled it.
    """
    from receipts.review.serializers import query_export_receipts

    statements: list[str] = []

    def _record(conn, cursor, statement, parameters, context, executemany):
        statements.append(statement)

    with session_factory() as session:
        engine = session.get_bind()
        event.listen(engine, "before_cursor_execute", _record)
        try:
            query_export_receipts(
                session, status=None, merchant_id=None, date_from=None,
                date_to=None, min_confidence=None, limit=1, offset=0,
            )
        finally:
            event.remove(engine, "before_cursor_execute", _record)

    ordered = [one for one in statements if "ORDER BY" in one]
    assert len(ordered) == 1, f"expected exactly one ordered SELECT, got {statements}"
    order_by = ordered[0][ordered[0].index("ORDER BY"):]

    assert "receipts.created_at" in order_by, f"created_at missing from: {order_by}"
    assert "receipts.id" in order_by, f"the id tie-break is missing from: {order_by}"
    assert order_by.index("receipts.created_at") < order_by.index("receipts.id")


def _listed_receipt_ids(client, **params) -> set[str]:
    """Every id ``GET /export/receipts`` yields, paged to exhaustion.

    ``limit=1`` is chosen against this fixture rather than as a generic "small
    page". ``admin_client`` seeds exactly two exportable receipts, so one row
    per page is the largest page that still leaves a second one to fetch: the
    unfiltered call below makes two requests and sees ``has_more`` ``True`` on
    the first and ``False`` on the second.

    ``limit=2`` did neither. Both rows fitted the first page, ``has_more`` was
    ``False`` on the first response, and the ``offset`` line never ran -- so
    the route's ``True`` branch was pinned by nothing here. Callers that narrow
    the set further (``status="pending"`` matches one row) still finish in a
    single page; the unfiltered call is the one that exercises the loop.
    """
    ids: set[str] = set()
    offset = 0
    while True:
        response = client.get(
            "/export/receipts", params={**params, "limit": 1, "offset": offset}
        )
        assert response.status_code == 200, response.text
        body = response.json()
        ids.update(str(row["id"]) for row in body["items"])
        if not body["has_more"]:
            return ids
        offset += 1


def test_the_list_and_the_workbook_name_the_same_receipts(admin_client, pending_receipt_id):
    """One predicate at both ends -- design section 2, and the reason this route exists.

    The list is a projection of the export's own query, so the two cannot
    disagree about scope. Stated as a set equality rather than an enumeration
    of statuses: an enumeration would need editing every time a
    ``ReceiptStatus`` member is added, and the enumeration is what goes stale.
    """
    listed = _listed_receipt_ids(admin_client)
    in_workbook = _receipt_ids_in(admin_client.get("/export/xlsx"))

    # Anti-vacuity, both halves. Two empty sets are equal.
    assert listed, "fixture must produce at least one exportable receipt"
    assert listed == in_workbook

    # And the equality is not trivially "everything": the excluded status is
    # absent from both, and present in both once it is asked for.
    assert str(pending_receipt_id) not in listed

    asked = _listed_receipt_ids(admin_client, status="pending")
    asked_workbook = _receipt_ids_in(
        admin_client.get("/export/xlsx", params={"status": "pending"})
    )
    assert str(pending_receipt_id) in asked
    assert asked == asked_workbook


def test_the_list_is_visible_to_a_reviewer_the_workbook_is_not(reviewer_client):
    """The two routes share a scope predicate and differ only in guard.

    Seeing the ledger and extracting it are different acts (design decision 3).
    Pinned rather than commented, because matching guards is exactly what a
    later reader would "tidy" these two into.
    """
    assert reviewer_client.get("/export/receipts").status_code == 200
    assert reviewer_client.get("/export/xlsx").status_code == 403


# --------------------------------------------------------------------------- #
# Fix round 1 (F1, F2, F3): a leaked temp file, an unbounded read, and an
# N+1 a docstring claimed did not exist.
# --------------------------------------------------------------------------- #


def test_export_with_a_malformed_range_header_leaves_no_temp_file(admin_client):
    """The original implementation wrote the workbook under
    ``tempfile.mkdtemp()`` and returned a ``FileResponse`` whose cleanup was
    a ``BackgroundTask``. In starlette 1.3.1, ``FileResponse.__call__``
    returns early for a malformed or unsatisfiable ``Range`` header --
    skipping ``await self.background()`` -- so a ``Range: bytes=abc`` left
    a complete financial workbook behind in the shared OS temp directory,
    permanently, once per request. The fix builds the response body fully
    in memory and deletes its own working temp directory synchronously
    (``finally``, before ``return``), so there is no deferred cleanup step
    left for any response path -- malformed ``Range``, an out-of-bounds
    one, or a client that disconnects mid-stream -- to skip.
    """
    before = set(glob.glob(os.path.join(tempfile.gettempdir(), "receipts-export-*")))
    response = admin_client.get("/export/xlsx", headers={"Range": "bytes=abc"})
    after = set(glob.glob(os.path.join(tempfile.gettempdir(), "receipts-export-*")))

    assert after - before == set()
    # A plain Response has no partial-content support: Range is ignored and
    # the whole (valid) workbook comes back, which is the accepted
    # trade-off for never leaking one -- see the route's docstring.
    assert response.status_code == 200
    load_workbook(io.BytesIO(response.content))  # does not raise


def test_upload_bounds_the_read_by_the_configured_limit(client_max_1mb, monkeypatch):
    """``await file.read()`` with no argument buffers the *entire* body
    into one ``bytes`` object before ``ingest_bytes`` ever gets a chance to
    reject it on size -- an allocation bounded only by whatever the client
    chooses to send, from any reviewer session or the machine key. The fix
    reads ``max_bytes + 1`` instead, so ``ingest_bytes`` -- spied on here --
    never sees more than one byte past the configured cap, regardless of
    how much larger the uploaded body actually is.
    """
    captured_sizes: list[int] = []
    original_ingest_bytes = api_module.ingest_bytes

    def _spy(data, *args, **kwargs):
        captured_sizes.append(len(data))
        return original_ingest_bytes(data, *args, **kwargs)

    monkeypatch.setattr(api_module, "ingest_bytes", _spy)

    max_bytes = 1 * 1024 * 1024  # client_max_1mb's configured cap
    big = b"\xff\xd8" + b"\x00" * (32 * 1024 * 1024)
    response = client_max_1mb.post("/upload", files={"file": ("r.jpg", big, "image/jpeg")})

    assert response.status_code == 400
    assert captured_sizes == [max_bytes + 1]


def test_export_does_not_issue_one_query_per_receipt_for_line_items_or_merchant(
    admin_client, session_factory
):
    """``build_export_rows`` touches ``receipt.line_items`` (to rebuild the
    extraction) and ``receipt.merchant`` (for the canonical name) for every
    row. Both are default-lazy relationships
    (:mod:`receipts.persist.models`); left lazy, each access is its own
    SELECT -- an N+1 a two-receipt test run does not surface but that
    becomes 5,000-10,000 round trips at the route's own
    ``_EXPORT_MAX_ROWS`` design point. ``_query_export_receipts`` now
    eager-loads both with ``selectinload``, so the statement count must not
    grow with the number of matching receipts.

    Two of the extra receipts get **distinct** ``Merchant`` rows, on
    purpose: every other receipt fixture in this module leaves
    ``merchant_id`` NULL, and a many-to-one lazy loader skips the query
    entirely when the local FK is NULL -- a NULL-only fixture would record
    zero merchant statements whether or not ``selectinload(Receipt.
    merchant)`` is present, discriminating on nothing (this is exactly the
    gap a re-review caught in the first version of this test). Two
    *distinct* merchants close it a second way too: SQLAlchemy's lazy
    loader checks the session's identity map before it queries, so if both
    receipts pointed at the *same* merchant, the second access could be
    served from the map with no second query even without eager loading.
    Distinct merchants force two genuinely separate per-object SELECTs when
    unbatched. The third extra receipt (plus ``receipt_id``/
    ``other_receipt_id`` from ``admin_client``'s own fixture) keeps
    ``merchant_id`` NULL, which also exercises the ``merchant is None``
    fallback in :func:`~receipts.review.serializers.build_export_rows`.
    """
    # admin_client's own fixture already seeded two receipts (receipt_id,
    # other_receipt_id, both merchant_id=NULL); add three more so a
    # per-receipt query pattern is unmistakable against a batched one.
    with session_factory() as session:
        merchant_a = Merchant(id=uuid.uuid4(), canonical_name="Merchant A", name_variants=[])
        merchant_b = Merchant(id=uuid.uuid4(), canonical_name="Merchant B", name_variants=[])
        session.add_all([merchant_a, merchant_b])
        session.flush()

        for merchant_id in (merchant_a.id, merchant_b.id, None):
            extra_id = uuid.uuid4()
            session.add(
                Receipt(
                    id=extra_id,
                    status=ReceiptStatus.AUTO_APPROVED,
                    confidence=Decimal("0.900"),
                    merchant_id=merchant_id,
                    merchant_name_raw="EXTRA CO",
                    txn_date=date(2026, 7, 3),
                    currency="USD",
                    total=Decimal("5.00"),
                    image_key=make_image_key(extra_id, "original"),
                    image_phash="",
                )
            )
        session.commit()
        engine = session.get_bind()

    statements: list[str] = []

    def _record(conn, cursor, statement, parameters, context, executemany):
        statements.append(statement)

    event.listen(engine, "before_cursor_execute", _record)
    try:
        response = admin_client.get("/export/xlsx")
    finally:
        event.remove(engine, "before_cursor_execute", _record)

    assert response.status_code == 200
    line_item_queries = [s for s in statements if "line_items" in s.lower()]
    merchant_queries = [s for s in statements if "merchants" in s.lower()]
    # Five qualifying receipts (two with distinct real merchants), at most
    # one batched SELECT each -- not five, and not two for merchant either.
    assert len(line_item_queries) <= 1
    assert len(merchant_queries) <= 1


def test_build_export_rows_without_a_secret_leaves_the_image_column_empty(
    session_factory, receipt_id
):
    from receipts.review.serializers import build_export_rows, query_export_receipts

    with session_factory() as session:
        receipts = query_export_receipts(
            session, status=None, merchant_id=None, date_from=None,
            date_to=None, min_confidence=None, limit=100,
        )
        _extractions, rows = build_export_rows(
            session, receipts, secret=None, image_url_ttl_s=86400
        )

    assert rows, "fixture should produce at least one exportable receipt"
    # An unverifiable link is worse than no link.
    assert all(row.image_url is None for row in rows)


def test_the_rebuilt_extraction_carries_the_buyer_and_the_template_flag(session_factory):
    """``_export_extraction`` is lossless for every column the export writes.

    The export consumes ``ReceiptExtraction``, not the ORM row (ADR-0010), so a
    column this rebuild drops is a column the spreadsheet can never show no
    matter what ``export/xlsx.py`` does. Both are asserted here rather than in
    the export's own tests because this is where the loss would happen: the
    buyer would export blank, and a template row -- a blank pre-printed line
    nobody bought -- would be indistinguishable from a purchase, which is an
    accounting ledger listing a sale that never occurred.
    """
    from receipts.persist.models import LineItem
    from receipts.review.serializers import build_export_rows, query_export_receipts

    receipt_uuid = uuid.uuid4()
    with session_factory() as session:
        session.add(
            Receipt(
                id=receipt_uuid,
                status=ReceiptStatus.AUTO_APPROVED,
                confidence=Decimal("0.910"),
                merchant_name_raw="TOTAL WINE",
                buyer_name_raw="IDEAL SOURCE",
                buyer_tax_id="123-456-789-000",
                txn_date=date(2026, 7, 4),
                currency="USD",
                total=Decimal("2000.00"),
                image_key=make_image_key(receipt_uuid, "original"),
                image_phash="",
                line_items=[
                    LineItem(position=0, description_raw="MaxiPower", is_template_row=True),
                    LineItem(
                        position=1,
                        description_raw="DieselPlus",
                        line_total=Decimal("2000.00"),
                    ),
                ],
            )
        )
        session.commit()

    with session_factory() as session:
        receipts = query_export_receipts(
            session, status=None, merchant_id=None, date_from=None,
            date_to=None, min_confidence=None, limit=100,
        )
        extractions, _rows = build_export_rows(
            session, receipts, secret="s", image_url_ttl_s=86400
        )

    # The query returns every exportable receipt in the shared database, so the
    # row this test seeded is picked out by its own merchant rather than by an
    # index another fixture could shift.
    matched = [e for e in extractions if e.merchant.name == "TOTAL WINE"]
    assert len(matched) == 1, f"expected exactly one seeded row, got {len(matched)}"
    extraction = matched[0]
    assert extraction.buyer.name == "IDEAL SOURCE"
    assert extraction.buyer.tax_id == "123-456-789-000"
    assert [item.is_template_row for item in extraction.line_items] == [True, False]


# --------------------------------------------------------------------------- #
# The 400 texts the review UI's failure classifier encodes (error-recovery
# milestone). The client matches quoted spans in these messages against the
# paths and values it just sent, so the exact wording is load-bearing on the
# other side of the wire. See
# docs/superpowers/specs/2026-08-03-review-ui-error-recovery-design.md §1.3.
# --------------------------------------------------------------------------- #


@pytest.mark.parametrize(
    ("body", "message"),
    [
        ({"totals.total": "abc"}, "not a decimal amount: 'abc'"),
        (
            {"totals.total": "nan"},
            "money must be a finite amount, not 'nan'; a non-finite value "
            "would destroy the stored amount and make the corrections audit "
            "row disagree with the column",
        ),
        (
            {"receipt.date": "14/07/2026"},
            "not an ISO 8601 date (YYYY-MM-DD): '14/07/2026'",
        ),
        ({"receipt.time": "2.30pm"}, "not an ISO 8601 time (HH:MM): '2.30pm'"),
        (
            {"receipt.currency": "EUROS"},
            "currency holds at most 3 characters, got 5 ('EUROS')",
        ),
    ],
)
def test_the_400_texts_the_client_matcher_encodes(
    reviewer_client, receipt_id, body, message
):
    """Each row is a value a reviewer can actually type into the UI.

    The value-coercion messages quote only the offending value, never the
    field path -- the classifier's value-quote rule exists because of that,
    so a wording change here must be mirrored in
    frontend/src/review/failure.ts and its tests.
    """
    response = reviewer_client.patch(f"/receipts/{receipt_id}", json=body)
    assert response.status_code == 400
    assert response.json() == {"error": {"message": message}}


def test_a_dotted_key_with_a_bad_value_is_the_valueerror_400_not_a_422(
    reviewer_client, receipt_id
):
    """The UI sends flat dotted keys, which bypass CorrectionPatch's typed
    sub-models (extra="allow", review/schemas.py:149) -- so even a JSON float
    smuggled under one reaches `_coerce_money` and comes back as the enveloped
    400, never FastAPI's 422 shape. The classifier's whole 400 orientation
    rests on this division."""
    response = reviewer_client.patch(f"/receipts/{receipt_id}", json={"totals.total": 1.5})
    assert response.status_code == 400
    body = response.json()
    assert "detail" not in body
    assert body["error"]["message"] == (
        "money must be a Decimal or a string, not float (1.5); "
        "a float cannot represent an exact amount"
    )


def test_logout_returns_204_with_an_empty_body_and_ends_the_session(
    reviewer_client, receipt_id
):
    """The SignOutControl's contract: 204, no body (client.ts resolves an
    empty body to `undefined`), and the session is really over."""
    response = reviewer_client.post("/auth/logout")
    assert response.status_code == 204
    assert response.content == b""
    after = reviewer_client.get(f"/receipts/{receipt_id}")
    assert after.status_code == 401


# The path-quoting family (the five above are the value-quoting family): these
# messages quote the *field path* the reviewer sent, which is what the
# classifier's path rule matches on.


def test_an_unknown_field_path_is_a_400_naming_the_path(reviewer_client, receipt_id):
    """``_plan_change`` refuses a path outside ``_RECEIPT_FIELDS`` rather than
    silently dropping the edit, and names the offending path in the message --
    the classifier reads that quoted span to blame the right input."""
    response = reviewer_client.patch(
        f"/receipts/{receipt_id}", json={"totals.grand_total": "1.00"}
    )
    assert response.status_code == 400
    assert response.json() == {
        "error": {
            "message": "cannot apply a correction to unknown field path 'totals.grand_total'"
        }
    }


def test_a_line_item_position_that_does_not_exist_is_a_400_naming_path_and_receipt(
    reviewer_client, receipt_id
):
    """The other half of the path family: a *known* line-item field at a
    position this receipt has no item for. The path is quoted, the receipt id
    and the position are bare -- the classifier's path rule must not assume
    every interpolated value in this family is quoted.
    """
    response = reviewer_client.patch(f"/receipts/{receipt_id}", json={"line_items[9].qty": "1"})
    assert response.status_code == 400
    assert response.json() == {
        "error": {
            "message": (
                "cannot apply a correction to 'line_items[9].qty': "
                f"receipt {receipt_id} has no line item at position 9"
            )
        }
    }


def test_a_pan_never_reaches_the_corrections_route(reviewer_client, receipt_id, task_id):
    """The fourth place, and it did not exist until the corrections route did.

    ``test_a_dotted_pan_is_masked_in_the_row_the_body_and_the_audit_copy`` reads
    ``corrections.value_after`` straight out of the database. That was the only
    way to reach it. This route serves the same column over HTTP, so the
    masking now has a network egress it never had, and the guarantee has to be
    asserted where a client actually sees it.

    Takes the existing ``task_id`` fixture for its **side effect**, not its
    value: it enqueues the seeded receipt and claims it for alice, which is what
    entitles her to read the history. Without it this test would get a 403 and
    assert nothing about redaction -- see Step 9, where that is exactly the
    wrong-reason failure to watch for.

    Goes red if ``_plan_change``'s ``after = redact_pan(after)`` is removed.
    """
    typed = "VISA 4111111111111111"
    assert reviewer_client.patch(
        f"/receipts/{receipt_id}", json={"payment": {"method": typed}}
    ).status_code == 200

    response = reviewer_client.get(f"/receipts/{receipt_id}/corrections")

    assert response.status_code == 200
    assert "4111111111111111" not in response.text
    assert typed not in response.text
    assert [row["value_after"] for row in response.json()["items"]] == ["VISA ************1111"]


def test_a_configured_api_key_cannot_read_a_correction_history(key_client, receipt_id):
    """The machine key authorizes upload and nothing else (spec section 5.3).

    ``test_corrections_require_a_session[api_key]`` in ``tests/test_api_read.py``
    cannot say this. That module never sets ``RECEIPTS_API_KEY``, so its
    ``api_key`` client carries a header no configured key matches and is
    indistinguishable from an anonymous one -- it pins "an unrecognised header
    is 401", not "the machine key may not read an audit trail". This module
    configures a real key, which is what makes the question askable.

    Measured, and the measurement is not what the mutant first looked like.
    With this route's dependency changed from ``require_user`` to
    ``require_upload`` -- the one ``POST /upload`` itself uses -- the whole suite
    stayed green as it stood at ``6536d0f``, the commit before this test existed;
    this test is what closes it, so that mutation is expected to fail *here*
    now. (Dated rather than counted: a bare suite total rots without its
    sentence changing, and read as present tense it would be false.) But a
    machine key does **not** reach a served history:
    ``require_upload`` returns ``None`` for a valid key (a machine, not a
    person, see its docstring), so ``user.role`` on the route's first line
    raises ``AttributeError: 'NoneType' object has no attribute 'role'`` and the
    caller gets a 500. What that mutant exposes is that **nothing in the suite
    told the two dependencies apart at all** -- not that one of them served the
    audit trail.

    Which is why this asserts the status code rather than the absence of a body:
    401 fails against the 500 that mutant produces today, and would equally fail
    against a 200 if some later change ever handed a machine key a real
    ``SessionUser``. Both are the same defect reaching the caller differently.
    """
    response = key_client.get(
        f"/receipts/{receipt_id}/corrections", headers={"X-API-Key": "s3cret-machine-key"}
    )

    assert response.status_code == 401
