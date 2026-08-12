"""``receipts export``, ``receipts merchants``, ``receipts eval`` and
``receipts calibrate`` (P4.T4/T5, spec 14.10 / §13 / §16 / §18).

Everything here is offline: a file-backed SQLite database and no storage
backend at all -- export reads only ORM rows, never a blob, through Task 1's
``query_export_receipts``/``build_export_rows``. No provider, no Redis, no
network. ``eval``'s tests inject a ``run_baseline_fn`` double rather than
touch the pipeline or monkeypatch a module global, and ``calibrate`` only
ever reads a hand-written results JSON -- neither needs a database or a
provider either.

The load-bearing behaviours pinned down below:

  * ``export`` writes the same four §13 sheets (``Receipts``, ``LineItems``,
    ``Needs Review``, ``Summary``) that ``GET /export/xlsx`` writes, because
    both call the identical ``query_export_receipts``/``build_export_rows``
    pair -- a CLI export and an API export of the same filters can never
    disagree about which receipts qualify.
  * ``pending``/``rejected`` receipts are excluded by default -- a pending row
    is an upload in flight, not a transaction -- unless ``--status`` names one
    of them explicitly.
  * Past ``_EXPORT_MAX_ROWS`` matching receipts, ``export`` refuses and writes
    nothing, rather than silently truncating what would read as a complete
    ledger.
  * ``export`` needs no ``SESSION_SECRET``: with none configured, the image
    column is simply empty rather than an unsigned or fabricated link.
  * ``merchants hints --add`` rebinds ``Merchant.hints`` rather than mutating
    it in place. The column is a plain ``sa.JSON()`` with no ``MutableList``
    registered, so an in-place ``.append()`` would silently never reach the
    database -- and the identity map would still hand the mutated object back
    within the *same* session, so a same-session assertion would pass while
    the database held nothing. Every hints test below re-reads through a new
    ``session_factory()`` block for exactly that reason; that shape is
    deliberate and must not be "simplified" away.
  * §18: a merchant hint always ends by deferring to the image -- ``--add``
    appends "; trust the image" when the supplied text does not already end
    with it (case-insensitively), and says on stdout that it did.
  * ``eval`` -- the *producer* -- refuses a zero-receipt run before it prints
    anything, and validates ``--golden-dir`` before the baseline runs at all
    so no results file is written. The guard used to sit only on
    ``calibrate``, the *reader*: a typo'd ``--golden-dir`` raised nothing
    (globbing a missing directory yields no labels), so ``eval`` printed
    ``Auto-approval precision: 100.00%`` over zero receipts, exited ``0``,
    and left a results file saying the same -- which then shadowed the
    genuine baseline in that directory, because ``latest_results_file``
    sorts by mtime.
  * ``calibrate`` refuses outright on a zero-receipt result set, printing no
    precision figure at all -- this project has already committed a results
    artifact reporting ``auto_approval_precision: 1.0`` on zero receipts
    once, and the command whose job is choosing an auto-approval threshold
    is the worst place to repeat it.
  * ``calibrate``'s recommendation is floored at ``REVIEW_THRESHOLD`` and
    needs ``_MIN_APPROVED_SAMPLE`` approved receipts behind it. The previous
    scan took the lowest threshold with a *nonzero* rate and sufficient
    precision, which on a fully-correct golden set is ``0`` -- auto-approve
    everything, at any confidence, forever -- and on a mostly-wrong one is
    whatever threshold a single lucky receipt cleared, reported as 100%.
    Both are the vacuous-precision trap: a number that is arithmetically
    true and evidentially worthless.
  * ``eval``/``calibrate`` import ``eval.*``, and every command that needs
    the optional ``pipeline`` extra imports *that*, lazily inside the
    function -- never at module top. Each one broke every ``receipts``
    command, not only its own, the instant the CLI ran somewhere the
    package was genuinely absent. ``pytest``'s own
    ``pythonpath = ["src", "."]`` hides the ``eval`` half in-process and a
    dev machine with every extra installed hides the other half, so an
    in-process assertion cannot pin either:
    ``test_cli_imports_without_the_eval_package_or_the_pipeline_extra``
    below runs in a subprocess with a ``sys.meta_path`` finder that blocks
    ``eval``/``eval.*`` and every name in ``cli._PIPELINE_EXTRA_MODULES``
    (plus ``numpy``, which reaches the CLI's import graph transitively
    through ``receipts.ingest.dedupe`` without being declared in the extra)
    before ``import receipts.cli``, the same technique
    ``tests/test_import_isolation.py`` uses for the FastAPI check.
"""

from __future__ import annotations

import json
import subprocess
import sys
import uuid
from decimal import Decimal as D
from pathlib import Path

import pytest
from openpyxl import load_workbook

from config.settings import Settings
from eval.metrics import EvalReport, FieldBreakdown
from eval.run_baseline import format_report
from receipts import cli as cli_module
from receipts.cli import (
    EXIT_FAILED,
    EXIT_OK,
    build_parser,
    cmd_calibrate,
    cmd_eval,
    cmd_export,
    cmd_merchants,
)
from receipts.extract.schema import LineItem as ExtractLineItem
from receipts.extract.schema import Merchant as ExtractMerchant
from receipts.extract.schema import ReceiptExtraction, ReceiptMeta, Totals
from receipts.ingest.ingest import ReceiptJob
from receipts.persist.models import Base, Merchant
from receipts.persist.repository import save_extraction
from receipts.persist.session import make_engine, make_session_factory
from receipts.score.confidence import ReceiptStatus
from receipts.score.thresholds import REVIEW_THRESHOLD
from receipts.validate.report import ValidationReport

# --------------------------------------------------------------------------- #
# Fixtures
# --------------------------------------------------------------------------- #


@pytest.fixture()
def session_factory(tmp_path):
    """A file-backed SQLite database, so several sessions share it."""
    engine = make_engine(f"sqlite:///{(tmp_path / 'receipts.db').as_posix()}")
    Base.metadata.create_all(engine)
    return make_session_factory(engine)


@pytest.fixture()
def settings() -> Settings:
    """A session secret set, so most tests below exercise the signed
    image-URL path in ``build_export_rows`` (the normal, configured case).
    ``test_export_without_a_session_secret_still_writes_a_workbook`` builds
    its own ``Settings`` with ``session_secret=None`` to prove the CLI needs
    no session at all.
    """
    return Settings(_env_file=None, session_secret="test-secret")


# --------------------------------------------------------------------------- #
# Test helpers
# --------------------------------------------------------------------------- #


def _receipt(session_factory, *, status: ReceiptStatus, total: D = D("224.00")) -> uuid.UUID:
    """One persisted receipt in the given terminal status."""
    job = ReceiptJob(
        id=uuid.uuid4(), image_key=f"receipts/2026/07/{uuid.uuid4()}/original.jpg",
        source="test", original_filename="r.jpg", content_type="image/jpeg",
    )
    extraction = ReceiptExtraction(
        merchant=ExtractMerchant(name="METRO OIL SUBIC, INC."),
        receipt=ReceiptMeta(date="2026-03-23", currency="PHP"),
        totals=Totals(total=total),
        line_items=[ExtractLineItem(position=1, description_raw="CLEAN DIESEL",
                                    line_total=total)],
    )
    with session_factory() as session:
        save_extraction(session, job, extraction, ValidationReport(), D("0.95"),
                        status, image_phash="a1b2c3d4a1b2c3d4")
        session.commit()
    return job.id


def _auto_approved_receipt(session_factory) -> uuid.UUID:
    return _receipt(session_factory, status=ReceiptStatus.AUTO_APPROVED)


def _pending_receipt_row(session_factory) -> uuid.UUID:
    return _receipt(session_factory, status=ReceiptStatus.PENDING)


def _merchant(session_factory, *, canonical_name: str, tax_id: str | None = None) -> uuid.UUID:
    """One merchants row. ``name_variants``/``hints`` default to [] via the ORM."""
    merchant = Merchant(id=uuid.uuid4(), canonical_name=canonical_name, tax_id=tax_id)
    with session_factory() as session:
        session.add(merchant)
        session.commit()
        return merchant.id


def _receipt_ids_in(path: Path) -> set[str]:
    """The `receipt_id` column of the Receipts sheet -- it is the first column."""
    sheet = load_workbook(path)["Receipts"]
    return {str(row[0].value) for row in sheet.iter_rows(min_row=2) if row[0].value}


# --------------------------------------------------------------------------- #
# export
# --------------------------------------------------------------------------- #


def test_export_writes_a_workbook_with_all_four_sheets(tmp_path, session_factory, settings):
    _auto_approved_receipt(session_factory)
    out = tmp_path / "book.xlsx"
    args = build_parser().parse_args(["export", "--out", str(out)])

    code = cmd_export(args, session_factory=session_factory, settings=settings)

    assert code == EXIT_OK
    assert load_workbook(out).sheetnames == ["Receipts", "LineItems", "Needs Review", "Summary"]


def test_export_excludes_pending_and_rejected_unless_asked(tmp_path, session_factory, settings):
    pending_id = _pending_receipt_row(session_factory)
    _auto_approved_receipt(session_factory)

    cmd_export(build_parser().parse_args(["export", "--out", str(tmp_path / "a.xlsx")]),
               session_factory=session_factory, settings=settings)
    default_ids = _receipt_ids_in(tmp_path / "a.xlsx")

    cmd_export(build_parser().parse_args(
        ["export", "--out", str(tmp_path / "b.xlsx"), "--status", "pending"]),
        session_factory=session_factory, settings=settings)
    asked_ids = _receipt_ids_in(tmp_path / "b.xlsx")

    # A pending row is an upload in flight, not a transaction.
    assert str(pending_id) not in default_ids
    assert str(pending_id) in asked_ids


def test_export_refuses_rather_than_truncating(
    tmp_path, session_factory, settings, monkeypatch, capsys
):
    for _ in range(2):
        _auto_approved_receipt(session_factory)
    monkeypatch.setattr(cli_module, "_EXPORT_MAX_ROWS", 1)
    out = tmp_path / "book.xlsx"

    code = cmd_export(build_parser().parse_args(["export", "--out", str(out)]),
                      session_factory=session_factory, settings=settings)

    # A silently shortened export reads as a complete ledger.
    assert code == EXIT_FAILED
    assert "narrow" in capsys.readouterr().err.lower()
    assert not out.exists()


def test_export_to_an_existing_directory_fails_cleanly(tmp_path, session_factory, settings, capsys):
    """`--out` pointed at a directory is a typo, not a bug.

    `export_workbook` handed the path straight to openpyxl's writer, which
    surfaced as a raw `PermissionError` traceback on Windows
    (`IsADirectoryError` elsewhere) -- while every other way this command
    declines is a message plus EXIT_FAILED.
    """
    _auto_approved_receipt(session_factory)
    out_dir = tmp_path / "exports"
    out_dir.mkdir()

    code = cmd_export(build_parser().parse_args(["export", "--out", str(out_dir)]),
                      session_factory=session_factory, settings=settings)

    assert code == EXIT_FAILED
    assert "directory" in capsys.readouterr().err
    assert out_dir.is_dir()  # untouched


def test_export_rejects_a_non_finite_min_confidence():
    """`--min-confidence nan` is a usage error, not a filter that matches nothing.

    `Decimal("nan")` does not raise `InvalidOperation`: it is a legal Decimal
    that compares false against every stored confidence, so the typo produced a
    valid, *empty* workbook and exit 0 -- indistinguishable from "there really
    were no such receipts". `repository._coerce_money` refuses the same values
    on the write side for the same reason.
    """
    for bad in ("nan", "-nan", "inf", "-Infinity"):
        with pytest.raises(SystemExit) as exc:
            build_parser().parse_args(["export", "--out", "x.xlsx", "--min-confidence", bad])
        assert exc.value.code == 2


def test_export_without_a_session_secret_still_writes_a_workbook(tmp_path, session_factory):
    _auto_approved_receipt(session_factory)
    out = tmp_path / "book.xlsx"

    code = cmd_export(build_parser().parse_args(["export", "--out", str(out)]),
                      session_factory=session_factory,
                      settings=Settings(_env_file=None, session_secret=None))

    # The CLI needs no session; the image column is simply empty.
    assert code == EXIT_OK and out.exists()


# --------------------------------------------------------------------------- #
# merchants
# --------------------------------------------------------------------------- #


def test_merchants_list_prints_each_merchant(session_factory, capsys):
    _merchant(session_factory, canonical_name="METRO OIL SUBIC, INC.", tax_id="221 193 789 09013")
    code = cmd_merchants(build_parser().parse_args(["merchants", "list"]),
                         session_factory=session_factory)
    assert code == EXIT_OK
    assert "METRO OIL SUBIC, INC." in capsys.readouterr().out


def test_merchants_list_does_not_print_a_confident_zero_receipt_count(session_factory, capsys):
    """`merchants.receipt_count` is never incremented before Phase 6.

    Every merchant reads back the column default, so the list printed a column
    of `0`s -- a wrong number stated confidently, which this system treats as
    strictly worse than a missing one, and one an operator would read as "this
    merchant has no receipts" rather than "nothing counts them yet".
    """
    _merchant(session_factory, canonical_name="METRO OIL SUBIC, INC.", tax_id="221 193 789 09013")

    cmd_merchants(build_parser().parse_args(["merchants", "list"]),
                  session_factory=session_factory)

    line = capsys.readouterr().out.strip()
    assert line.split("\t")[-1] == cli_module._RECEIPT_COUNT_NOT_TRACKED
    assert line.split("\t")[-1] != "0"


def test_merchants_hints_add_appends_trust_the_image(session_factory):
    merchant_id = _merchant(session_factory, canonical_name="SUMMIT FUEL OPC")
    args = build_parser().parse_args(
        ["merchants", "hints", str(merchant_id), "--add", "The TIN is in the header"])

    cmd_merchants(args, session_factory=session_factory)

    with session_factory() as session:
        hints = session.get(Merchant, merchant_id).hints
    # §18: a merchant hint always ends by deferring to the image.
    assert hints[0].endswith("trust the image")


def test_merchants_hints_does_not_double_append(session_factory):
    merchant_id = _merchant(session_factory, canonical_name="SERV CENTRAL, INC.")
    args = build_parser().parse_args(
        ["merchants", "hints", str(merchant_id), "--add", "Read the header; trust the image"])

    cmd_merchants(args, session_factory=session_factory)

    with session_factory() as session:
        assert session.get(Merchant, merchant_id).hints[0].count("trust the image") == 1


def test_merchants_hints_for_an_unknown_id_is_exit_one(session_factory):
    args = build_parser().parse_args(["merchants", "hints", str(uuid.uuid4())])
    assert cmd_merchants(args, session_factory=session_factory) == EXIT_FAILED


# --------------------------------------------------------------------------- #
# eval / calibrate test helpers
# --------------------------------------------------------------------------- #


def _write_results(results_dir: Path, *, receipts: int, results: list[dict]) -> Path:
    """A results file in the exact shape `eval/harness.py::_report_to_dict` writes."""
    results_dir.mkdir(parents=True, exist_ok=True)
    path = results_dir / "2026-07-29-1.0.0.json"
    path.write_text(json.dumps({
        "prompt_version": "1.0.0",
        "auto_approve_threshold": "0.85",
        "counts": {"receipts": receipts, "auto_approved": 0,
                   "critical_correct": 0, "failed": 0},
        "metrics": {"auto_approval_precision": 0.0, "auto_approval_rate": 0.0,
                    "critical_field_accuracy": 0.0, "transcription_accuracy": 0.0,
                    "transcription_accuracy_core": 0.0,
                    "transcription_accuracy_line_items": 0.0,
                    "self_report_agreement": 0.0, "hallucinated_fields": 0,
                    "correctly_empty_fields": 0,
                    "line_item_precision": 0.0, "line_item_recall": 0.0,
                    "line_item_f1": 0.0, "cost_per_receipt": None,
                    "p50_latency_s": None, "p95_latency_s": None},
        "failures": [],
        "calibration": [],
        "results": results,
    }, indent=2), encoding="utf-8")
    return path


def _stub_report() -> EvalReport:
    """Just enough EvalReport for `format_report` to render the table."""
    return EvalReport(
        n_receipts=3, n_auto_approved=2, n_critical_correct=2,
        auto_approve_threshold=D("0.85"),
        auto_approval_precision=1.0, auto_approval_rate=0.667,
        critical_field_accuracy=0.667,
        # 17/20 == the 0.85 this fixture carried as a single scalar. The counts
        # are the stored form now; the ratios are derived from them.
        breakdown=FieldBreakdown(transcription_correct=17, transcription_total=20,
                                 core_correct=17, core_total=20),
        line_item_precision=0.9, line_item_recall=0.9, line_item_f1=0.9,
    )


def _empty_report() -> EvalReport:
    """What `run_eval` returns for a golden set with no labels in it.

    Not invented: `_build_report` really does define `auto_approval_precision`
    as `1.0` when nothing was approved (`eval/harness.py:104`), and nothing
    approves anything when nothing was scored.
    """
    return EvalReport(
        n_receipts=0, n_auto_approved=0, n_critical_correct=0,
        auto_approve_threshold=D("0.85"),
        auto_approval_precision=1.0, auto_approval_rate=0.0,
        critical_field_accuracy=0.0,
        # Nothing scored, so every count is zero -- and every derived ratio is
        # `None`, not a `0.0` that would read as "measured, and bad".
        breakdown=FieldBreakdown(),
        line_item_precision=0.0, line_item_recall=0.0, line_item_f1=0.0,
    )


def _golden_dir(root: Path, *, labels: int = 1) -> Path:
    """A golden-set directory shaped the way `run_eval` expects to find one."""
    golden = root / "golden"
    (golden / "labels").mkdir(parents=True)
    (golden / "images").mkdir(parents=True)
    for index in range(labels):
        (golden / "labels" / f"r{index:03d}.json").write_text("{}", encoding="utf-8")
    return golden


def _recommended_threshold(out: str) -> str | None:
    """The value on `calibrate`'s recommendation line, or ``None``.

    Parsed out explicitly rather than asserted with a substring test: the
    printed curve contains a row for every candidate threshold, so
    ``assert "0.9" in out`` passes whatever the command actually recommends --
    including `0`, `1.0`, or nothing at all. That is why Critical 2 survived a
    test named for the recommendation.
    """
    for line in out.splitlines():
        if line.startswith("recommended threshold:"):
            return line.split(":", 1)[1].strip().split()[0]
    return None


# --------------------------------------------------------------------------- #
# eval
# --------------------------------------------------------------------------- #


def test_eval_prints_the_six_metric_table(capsys):
    report = _stub_report()
    code = cmd_eval(
        build_parser().parse_args(["eval"]),
        settings=Settings(_env_file=None),
        run_baseline_fn=lambda **kw: report,
    )

    assert code == EXIT_OK
    out = capsys.readouterr().out
    # These are `format_report`'s real labels, verified against
    # eval/run_baseline.py -- it prints prose headings, NOT the snake_case
    # metric names. Assert on what the function actually emits.
    for label in ("Auto-approval precision:", "Critical-field accuracy:",
                  "Line-item F1:", "Cost per receipt:"):
        assert label in out


def test_eval_reports_a_refused_provider_as_exit_one(capsys):
    def refuse(**kwargs):
        raise RuntimeError("the fake provider carries no scripted responses")

    code = cmd_eval(
        build_parser().parse_args(["eval"]),
        settings=Settings(_env_file=None),
        run_baseline_fn=refuse,
    )

    assert code == EXIT_FAILED
    assert "scripted responses" in capsys.readouterr().err


def test_eval_refuses_a_zero_receipt_run(capsys):
    """`eval` must never print a metric for a run that scored no receipts.

    `_build_report` defines auto-approval precision as `1.0` when nothing was
    approved, and nothing is approved when nothing was scored -- so a run over
    an empty golden set printed `Auto-approval precision: 100.00%` to an
    operator's terminal and exited 0. That is the exact artifact this project
    committed once and formally banned; the guard was put on `calibrate`, the
    reader, and never on `eval`, the producer.
    """
    code = cmd_eval(
        build_parser().parse_args(["eval"]),
        settings=Settings(_env_file=None),
        run_baseline_fn=lambda **kw: _empty_report(),
    )

    captured = capsys.readouterr()
    assert code == EXIT_FAILED
    assert "zero receipts" in captured.err
    # No precision figure at all -- not the vacuous 100%, not any other.
    assert "100.00%" not in captured.out
    assert "Auto-approval precision" not in captured.out


def test_eval_refuses_a_nonexistent_golden_dir_before_running_anything(tmp_path, capsys):
    """A typo'd `--golden-dir` raises nothing on its own: globbing a directory
    that does not exist simply yields no labels.

    The check has to happen *before* `run_baseline`, not after: the results
    file is written inside `run_eval`, so by the time a report comes back it is
    already on disk -- and `latest_results_file` sorts by mtime, so a poisoned
    zero-receipt file becomes the newest and makes the next bare `receipts
    calibrate` refuse with "zero receipts" instead of reading the genuine
    earlier baseline sitting beside it. One typo, calibration poisoned until a
    human deletes the file.
    """
    called = []

    def must_not_run(**kwargs):
        called.append(kwargs)
        return _stub_report()

    missing = tmp_path / "goldn"  # the typo
    code = cmd_eval(
        build_parser().parse_args(["eval", "--golden-dir", str(missing)]),
        settings=Settings(_env_file=None),
        run_baseline_fn=must_not_run,
    )

    assert code == EXIT_FAILED
    assert called == [], "the baseline ran, so a results file was already written"
    assert str(missing) in capsys.readouterr().err


def test_eval_refuses_a_golden_dir_with_no_labels(tmp_path, capsys):
    """The empty-directory half of the same defect -- and, per ADR-0013, the
    shape the banned artifact was actually produced from.
    """
    golden = _golden_dir(tmp_path, labels=0)
    called = []

    code = cmd_eval(
        build_parser().parse_args(["eval", "--golden-dir", str(golden)]),
        settings=Settings(_env_file=None),
        run_baseline_fn=lambda **kw: called.append(kw) or _stub_report(),
    )

    assert code == EXIT_FAILED
    assert called == []
    assert "no labels" in capsys.readouterr().err


def test_eval_runs_against_a_valid_golden_dir(tmp_path, capsys):
    """The guard above must not refuse a golden set that is genuinely fine."""
    golden = _golden_dir(tmp_path, labels=2)

    code = cmd_eval(
        build_parser().parse_args(["eval", "--golden-dir", str(golden)]),
        settings=Settings(_env_file=None),
        run_baseline_fn=lambda **kw: _stub_report(),
    )

    assert code == EXIT_OK
    assert "Auto-approval precision:" in capsys.readouterr().out


def test_format_report_never_prints_a_vacuous_hundred_percent_precision():
    """Belt to `cmd_eval`'s braces: no *caller* of `format_report` can print a
    100% precision that rests on nothing.

    `EvalReport.auto_approval_precision` is `1.0` whenever `n_auto_approved`
    is 0 -- "of the receipts we approved, all were correct" is vacuously true
    of an empty set -- and this is the line an operator reads as the system's
    headline metric.
    """
    approved_nothing = _empty_report()
    rendered = format_report(approved_nothing)

    assert "Auto-approval precision:" in rendered
    assert "100.00%" not in rendered
    assert "n/a" in rendered
    # A report that did approve something still prints the real number.
    assert "100.00%" in format_report(_stub_report())


def test_cli_imports_without_the_eval_package_or_the_pipeline_extra():
    """`receipts.cli` (and, transitively, every `receipts` command) must not
    require the `eval` package *or* the optional `pipeline` extra to import.

    `eval/` is deliberately excluded from the installed distribution
    (pyproject.toml: dev/research tooling, not part of the installed CLI), and
    `openpyxl`/`pillow`/`opencv-python-headless` are the `pipeline` extra --
    the base distribution installs only alembic/pydantic/pydantic-settings/
    pyyaml/sqlalchemy. A module-top import of either breaks every command the
    instant that is actually true: `receipts users list`, `receipts users add`,
    `receipts merchants list` and `receipts ingest` need neither a spreadsheet
    writer nor an image library, and all four died before argparse ran.
    Reproduced against the real installed console script from outside this
    repository, for `eval` and then again for `openpyxl`.

    Run in a subprocess with a `sys.meta_path` finder that raises
    `ModuleNotFoundError` for each blocked root package, installed before
    `import receipts.cli` -- deterministic and platform-independent, unlike
    relying on cwd/pythonpath tricks. An in-process assertion cannot pin
    either half: pytest's own `pythonpath = ["src", "."]` puts the repo root on
    `sys.path` (hiding the `eval` case) and a dev machine has every extra
    installed (hiding this one). Same technique as
    `tests/test_import_isolation.py`'s FastAPI check, adapted to block an
    import outright rather than inspect `sys.modules` after the fact.

    The blocked set is derived from `cli_module._PIPELINE_EXTRA_MODULES`
    itself, unioned with `numpy` (which is not declared there -- it arrives
    transitively through opencv/Pillow -- but is imported directly by
    `receipts.ingest.dedupe`), rather than restating a hand-picked subset here.
    A hand-picked `{'PIL', 'openpyxl', 'cv2'}` is what this test used to block:
    `_PIPELINE_EXTRA_MODULES` also names `pillow_heif` and `pypdfium2`, so a
    module-top import of either would have broken every installed `receipts`
    command while this test stayed green. Deriving from the real constant
    means the guard and the extra it mirrors cannot drift apart again.
    """
    blocked = sorted(cli_module._PIPELINE_EXTRA_MODULES | {"numpy"})
    code = (
        "import sys\n"
        f"_BLOCKED = set({blocked!r})\n"
        "class _BlockOptional:\n"
        "    def find_spec(self, name, path=None, target=None):\n"
        "        root = name.split('.')[0]\n"
        "        if root in _BLOCKED or name == 'eval' or name.startswith('eval.'):\n"
        "            raise ModuleNotFoundError(f'No module named {name!r}', name=name)\n"
        "        return None\n"
        "sys.meta_path.insert(0, _BlockOptional())\n"
        "import receipts.cli\n"
        "receipts.cli.build_parser()\n"
        "print('OK')\n"
    )
    result = subprocess.run(
        [sys.executable, "-c", code], capture_output=True, text=True, timeout=120
    )
    assert result.returncode == 0, result.stderr
    assert "OK" in result.stdout


def test_is_missing_eval_only_matches_the_eval_package():
    """`cmd_eval`/`cmd_calibrate` re-raise anything this rejects.

    A regression to a bare `except ModuleNotFoundError` would tell an operator
    running from a perfectly good checkout to "run this from a checkout of the
    repository" when the real problem is a missing `pipeline` extra several
    frames deeper inside `eval.run_baseline`'s own import of
    `receipts.pipeline` -- a confident, specific, wrong diagnosis.
    """
    assert cli_module._is_missing_eval(ModuleNotFoundError("x", name="eval"))
    assert cli_module._is_missing_eval(ModuleNotFoundError("x", name="eval.run_baseline"))

    assert not cli_module._is_missing_eval(ModuleNotFoundError("x", name="PIL"))
    assert not cli_module._is_missing_eval(ModuleNotFoundError("x", name="receipts.pipeline"))
    # A package whose name merely starts with "eval" is not `eval`.
    assert not cli_module._is_missing_eval(ModuleNotFoundError("x", name="evaluate"))
    # `.name` is optional on ModuleNotFoundError and must not blow up here.
    assert not cli_module._is_missing_eval(ModuleNotFoundError("x"))


# --------------------------------------------------------------------------- #
# calibrate
# --------------------------------------------------------------------------- #


def test_calibrate_refuses_a_zero_receipt_result_set(tmp_path, capsys):
    _write_results(tmp_path, receipts=0, results=[])
    code = cmd_calibrate(build_parser().parse_args(["calibrate"]), results_dir=tmp_path)

    # This project has already produced a 0/0 precision of 1.0 once. The command
    # whose job is choosing an auto-approval threshold is the worst place to
    # repeat it.
    assert code == EXIT_FAILED
    err = capsys.readouterr().err
    assert "zero receipts" in err.lower()
    assert "1.0" not in err


def _results(*specs: tuple[str, bool], count: int = 1) -> list[dict]:
    """``(confidence, critical_correct)`` pairs, each repeated ``count`` times."""
    rows: list[dict] = []
    for confidence, correct in specs:
        for _ in range(count):
            rows.append({
                "receipt_id": f"r{len(rows):03d}", "confidence": confidence,
                "critical_correct": correct,
                "transcription_correct": int(correct), "transcription_total": 1,
                "self_report_correct": 0, "self_report_total": 0,
                "hallucinated": 0, "correctly_empty": 0,
                "field_results": {},
            })
    return rows


def test_calibrate_recommends_the_lowest_threshold_clearing_the_target(tmp_path, capsys):
    """Six receipts at 0.90 are all critical-correct; four at 0.70 are not.

    0.70 admits the incorrect four, so precision there is 0.60; 0.8 admits only
    the six correct ones. The recommendation is asserted **explicitly**: the
    predecessor of this test asserted `"0.9" in out` while the command actually
    recommended `0.8`, and passed only because the printed curve happens to
    contain a `0.90` row -- so a regression recommending the *highest* clearing
    threshold, or `1.0`, or `0`, passed it identically. That is a large part of
    why Critical 2 got through.
    """
    _write_results(tmp_path, receipts=10, results=(
        _results(("0.90", True), count=6) + _results(("0.70", False), count=4)
    ))
    code = cmd_calibrate(
        build_parser().parse_args(["calibrate", "--target", "0.99"]), results_dir=tmp_path)

    out = capsys.readouterr().out
    assert code == EXIT_OK
    assert _recommended_threshold(out) == "0.8"


def test_calibrate_never_recommends_a_threshold_that_approves_everything(tmp_path, capsys):
    """A golden set on which every receipt is critical-correct -- the expected
    outcome of a first clean baseline on a small curated set -- used to yield
    `recommended threshold: 0`.

    `calibration_curve`'s sweep starts at `0`, where everything is approved and
    precision is therefore a perfect 1.0. `Settings.auto_approve_threshold` has
    no lower bound and `route()` approves on `confidence >= auto_threshold`, so
    an operator who followed that recommendation would auto-approve every
    receipt at any confidence and no receipt would ever reach a human again.
    The command whose whole job is protecting auto-approval precision was
    recommending that the gate be switched off.

    The original patch here guarded the fail-*safe* direction (a threshold that
    approves nothing) and left the fail-*dangerous* one wide open.
    """
    _write_results(tmp_path, receipts=8, results=(
        _results(("0.95", True), count=5) + _results(("0.20", True), count=3)
    ))
    code = cmd_calibrate(
        build_parser().parse_args(["calibrate", "--target", "0.99"]), results_dir=tmp_path)

    out = capsys.readouterr().out
    recommended = _recommended_threshold(out)
    assert code == EXIT_OK
    assert recommended is not None
    assert D(recommended) >= REVIEW_THRESHOLD
    # Not the sweep's floor, where everything is approved at any confidence.
    assert D(recommended) > D("0")
    assert recommended == "0.6"


def test_calibrate_will_not_recommend_from_a_single_approved_receipt(tmp_path, capsys):
    """`rate > 0.0` excluded an approved set of size *zero*, never one of size
    *one*.

    Twenty receipts, one of which clears: the curve reports 100.00% precision
    at every threshold above the rest of the pack, on a sample of one, and the
    closing caveat cited the twenty. Every candidate threshold here is already
    above the review-threshold floor, so this isolates the sample floor.
    """
    _write_results(tmp_path, receipts=20, results=(
        _results(("0.95", True), count=1) + _results(("0.70", False), count=19)
    ))
    code = cmd_calibrate(
        build_parser().parse_args(["calibrate", "--target", "0.99"]), results_dir=tmp_path)

    out = capsys.readouterr().out
    assert code == EXIT_FAILED
    assert _recommended_threshold(out) is None
    assert "no threshold" in out


def test_calibrate_prints_the_sample_behind_every_precision_figure(tmp_path, capsys):
    """A precision of 100.00% means one thing over 200 receipts and another
    over one. The curve must show which.
    """
    _write_results(tmp_path, receipts=6, results=(
        _results(("0.90", True), count=5) + _results(("0.10", False), count=1)
    ))
    cmd_calibrate(build_parser().parse_args(["calibrate"]), results_dir=tmp_path)

    lines = capsys.readouterr().out.splitlines()
    header = next(line for line in lines if "threshold" in line and "precision" in line)
    assert "approved" in header and "correct" in header
    # The 0.8 row approves the five correct receipts and nothing else.
    row = next(line for line in lines if line.split()[:1] == ["0.8"])
    assert row.split()[-2:] == ["5", "5"]


def test_calibrate_when_no_threshold_clears_the_target_recommends_nothing(tmp_path, capsys):
    _write_results(tmp_path, receipts=2, results=[
        {"receipt_id": "r001", "confidence": "0.90", "critical_correct": False,
         "transcription_correct": 0, "transcription_total": 1,
         "self_report_correct": 0, "self_report_total": 0,
         "hallucinated": 0, "correctly_empty": 0, "field_results": {}},
        {"receipt_id": "r002", "confidence": "0.95", "critical_correct": False,
         "transcription_correct": 0, "transcription_total": 1,
         "self_report_correct": 0, "self_report_total": 0,
         "hallucinated": 0, "correctly_empty": 0, "field_results": {}},
    ])
    code = cmd_calibrate(
        build_parser().parse_args(["calibrate", "--target", "0.99"]), results_dir=tmp_path)

    out = capsys.readouterr().out.lower()
    assert code == EXIT_FAILED
    # Never return the least-bad number as though it passed.
    assert "no threshold" in out or "none" in out


def test_calibrate_with_no_results_file_is_exit_one(tmp_path, capsys):
    # Wrapped to stay under the project's 100-column limit; identical call/assertion
    # to the brief's verbatim single-line form.
    code = cmd_calibrate(build_parser().parse_args(["calibrate"]), results_dir=tmp_path)
    assert code == EXIT_FAILED
    assert "receipts eval" in capsys.readouterr().err


@pytest.mark.parametrize(
    ("name", "payload"),
    [
        # A receipt entry missing `critical_correct` -> KeyError.
        ("missing_key", {"results": [{"receipt_id": "r001", "confidence": "0.9"}]}),
        # `"confidence": null` -> TypeError out of Decimal(None).
        ("null_confidence", {"results": [
            {"receipt_id": "r001", "confidence": None, "critical_correct": True}]}),
        # `results` is not a list -> TypeError iterating a dict's keys as dicts.
        ("results_not_a_list", {"results": {"r001": "0.9"}}),
        # An entry that is not an object at all -> TypeError indexing a str.
        ("entry_not_an_object", {"results": ["r001"]}),
        # The whole document is a list, not an object -> AttributeError on .get.
        ("top_level_not_an_object", ["r001"]),
    ],
)
def test_calibrate_on_a_results_file_it_cannot_read_fails_cleanly(tmp_path, capsys, name, payload):
    """Every other error path in this command is a clean message plus
    EXIT_FAILED; these were raw tracebacks.

    A results file that is not quite the shape this version expects is an
    ordinary thing to hit -- an older run, a hand-edited file, a half-written
    one -- and there is nothing an operator can do with a `KeyError:
    'critical_correct'` stack trace out of a list comprehension.
    """
    path = tmp_path / "2026-07-29-1.0.0.json"
    path.write_text(json.dumps(payload), encoding="utf-8")

    code = cmd_calibrate(
        build_parser().parse_args(["calibrate", "--results", str(path)]), results_dir=tmp_path)

    assert code == EXIT_FAILED, name
    assert "error:" in capsys.readouterr().err


def test_calibrate_on_a_malformed_json_file_fails_cleanly(tmp_path, capsys):
    path = tmp_path / "2026-07-29-1.0.0.json"
    path.write_text("{not json at all", encoding="utf-8")

    code = cmd_calibrate(
        build_parser().parse_args(["calibrate", "--results", str(path)]), results_dir=tmp_path)

    assert code == EXIT_FAILED
    assert "not valid JSON" in capsys.readouterr().err
